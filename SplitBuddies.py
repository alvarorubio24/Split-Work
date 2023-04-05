
from email import header
from genericpath import exists
from multiprocessing.connection import answer_challenge
from posixpath import split
import tkinter 
import tkinter.messagebox
import tkinter
import csv

from tkinter import ttk
from tkinter import scrolledtext
from tkinter.scrolledtext import ScrolledText
import tkinter as tk
from turtle import back, bgcolor
import webbrowser

import requests

import itertools

import openpyxl

import time
import os
import copy
import xlrd
import datetime
import time
import os.path
from os import path
from os import listdir
from os.path import isfile, join
from tkinter import *
from PIL import ImageTk, Image
import selenium
import csv
import pyperclip

def SplitBuddiesExecutable():
    #Open app zoomed
    root = tk.Tk()
    root.state('zoomed') 
    root.title('Split DS Buddies')
    root.iconbitmap(r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\PM_Dashboard\BuddysDSsplit\photoapp.ico')

    #scrollableframe
    class ScrollableFrame(ttk.Frame):
        def __init__(self, container, *args, **kwargs):
            super().__init__(container, *args, **kwargs)
            canvas = tk.Canvas(self)
            scrollbar = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
            Xscrollbar = ttk.Scrollbar(self, orient="horizontal", command=canvas.xview)
            self.scrollable_frame = ttk.Frame(canvas)

            self.scrollable_frame.bind(
                "<Configure>",
                lambda e: canvas.configure(
                    scrollregion=canvas.bbox("all")
                )
            )
            canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set, xscrollcommand = Xscrollbar.set)
            scrollbar.pack(side="right", fill="y")
            Xscrollbar.pack( side="bottom",  fill="x") # anchor = "se" )
            canvas.pack(side="left", fill="both",expand=True)

    frame = ScrollableFrame(root)

    # take user login
    user_login = os.getenv("username")


    #print(user_login)
    #user_login1 = "saffourm" #testing
    #user_login = "ledwidr"
        #mmordasz
        #wallrsa

    #Timeentry


    dateVar = datetime.datetime.today()
    dateStr1 = dateVar.strftime('%d-%m-%Y')
    dateStr2 = dateVar.strftime('%Y-%m-%d')
    tasklistlocation = r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\TaskList\TasklistGeneration\\' + dateStr2 + '\\TaskList ' + dateStr1 + '.xlsx'
    wb = openpyxl.load_workbook(tasklistlocation) #load workbook and telling Python to read only
    search_str = user_login

    def update(Timeentry):
        global Time
        global timeupdate
        Time = Timeentry.get()  #this is the entry the scheduler input to point what time he/she wants the break
        Timeentry.delete(0,END)
        tkinter.messagebox.showwarning("IMPORTANT","Make sure your break time is not clashing with the rest. \n To get the latest information click on Refresh button")
        timeupdate = datetime.datetime.now().strftime("%H:%M:%S")
        print(timeupdate)

        
        

        
        try: #updating excel file to add break timing to tasklist excel file
            os.chdir(r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\TaskList\TasklistGeneration\\' + dateStr2 )
            f = open("breakinfo.csv","a", newline="")
            tup = (NameScheduler, Time, timeupdate)
            writer = csv.writer(f)
            writer.writerow(tup)
            f.close()
            

        except Exception as FILEOPEN:
            print("File appears to be open. Please close the file and try again!")
            pass



    def refresh():   
        #workbook location
        dateVar = datetime.datetime.today()
        dateStr1 = dateVar.strftime('%d-%m-%Y')
        dateStr2 = dateVar.strftime('%Y-%m-%d')
        tasklistlocation = r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\TaskList\TasklistGeneration\\' + dateStr2 + '\\TaskList ' + dateStr1 + '.xlsx'
        wb = openpyxl.load_workbook(tasklistlocation) #load workbook and telling Python to read only
        search_str = user_login
        

        

        #Get Full Scheduler's name
        ws = wb.worksheets[0] #sheet number 1
        range = ws.iter_rows()
        for row in range:
                for cell in row:
                    if (cell.value == search_str):
                        global NameScheduler
                        NameScheduler =  ws.cell(row=cell.row, column=2).value #Full scheduler's name

        #creating the files we need
        os.chdir(r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\TaskList\TasklistGeneration\\' + dateStr2 )
        f = open("extracover.csv","a", newline="")
        f.close
        os.chdir(r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\TaskList\TasklistGeneration\\' + dateStr2 )
        f = open("breakinfo.csv","a", newline="")
        f.close
        os.chdir(r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\TaskList\TasklistGeneration\\' + dateStr2 )
        f = open("breakinfoconfirmation.csv","a", newline="")
        f.close
        
        #Get the buddies
        wBu = wb.worksheets[2]
        rangeBuddies = wBu.iter_rows()
        search_str2 = NameScheduler
        Buddy1list = []
        Buddy2list = []
        Buddy3list = []
        Buddy4list = []
        for row in rangeBuddies:
                for cell in row:
                    if (cell.value == search_str2):
                        global Buddy1
                        global Buddy2
                        global Buddy3
                        global Buddy4
                        Buddy1 = wBu.cell(row=cell.row, column=1).value
                        Buddy2 = wBu.cell(row=cell.row, column=2).value
                        Buddy3 = wBu.cell(row=cell.row, column=3).value
                        Buddy4 = wBu.cell(row=cell.row, column=4).value

                        Buddy1list.append(Buddy1)
                        Buddy2list.append(Buddy2)
                        Buddy3list.append(Buddy3)
                        Buddy4list.append(Buddy4)

                        print(Buddy1list[0])
                        print(Buddy2list[0])
                        print(Buddy3list[0])
                        print(Buddy4list[0])
        
        chimemessage = ((str(Buddy1) + " " + str(Buddy2) + " " + str(Buddy3) + " " + str(Buddy4)).replace(NameScheduler,"")).split("None")
        print(chimemessage)
        pyperclip.copy(chimemessage[0])
        
        #checking if 2,3 or 4 Buddy System
        #############################################2 buddy system####################################################
        ###############################################################################################################
        if Buddy3 == None and Buddy4==None:
            print("2 Buddy system")
            Buddy1 = Buddy1list[1] #take the 2nd element from the buddylist
            Buddy2 = Buddy2list[1] #do the same
            

            #creating frames for each buddy (3 frames in total)
            framebuddy1 = ttk.LabelFrame(frame.scrollable_frame, padding=5)
            framebuddy1.grid(row=1, column = 0, columnspan=3, \
                padx=5, pady=5, ipadx=5, ipady=5, sticky=N)
            
            framebuddy2 = ttk.LabelFrame(frame.scrollable_frame, padding=5)
            framebuddy2.grid(row=1, column = 4, columnspan=3, \
                padx=5, pady=5, ipadx=5, ipady=5,sticky=N)



            textlabelbuddy1 = Label(framebuddy1,  text= "Break for " + Buddy1, justify=CENTER, padx=20, pady=5, font=("Helvetica", 12, 'bold'))
            textlabelbuddy1.grid(row=2, column=0, columnspan=3, pady=10, padx=10)
            textlabelbuddy1cover1 = Label(framebuddy1,  text= Buddy2,justify=CENTER, padx=20, pady=5)
            textlabelbuddy1cover1.grid(row=4, column=0, columnspan=3, pady=10, padx=10)

            textlabelbuddy2 = Label(framebuddy2,  text= "Break for " + Buddy2 , justify=CENTER, padx=20, pady=5, font=("Helvetica", 12, 'bold'))
            textlabelbuddy2.grid(row=2, column=4, columnspan=3, pady=10, padx=10)
            textlabelbuddy2cover1 = Label(framebuddy2,  text= Buddy1, padx=20,justify=CENTER, pady=5)
            textlabelbuddy2cover1.grid(row=4, column=4, columnspan=3, pady=10, padx=10)


            
            ######################################################################################################################################
            #searching for the DS of the first buddy
            search_strbuddy1 = Buddy1
            search_strbuddy2 = Buddy2
            
            #getting the DS for 1st buddy and appending to a list
            DSscheduler1coverbuddy1 = []
            DSscheduler2coverbuddy1 = []
            ws = wb.worksheets[0] #sheet number 1
            range = ws.iter_rows()
            for row in range:
                    for cell in row:
                        if (cell.value == search_strbuddy1):
                            DSnode1buddy1 = ws.cell(row=cell.row, column=4).value #get the value from column 4 and the row where DS has been found
                            DSnode2buddy1 = ws.cell(row=cell.row, column=5).value
                            DSnode3buddy1 = ws.cell(row=cell.row, column=6).value
                            DSnode4buddy1 = ws.cell(row=cell.row, column=7).value
                            DSnode5buddy1 = ws.cell(row=cell.row, column=8).value
                            DSnode6buddy1 = ws.cell(row=cell.row, column=9).value
                            DSnode7buddy1 = ws.cell(row=cell.row, column=10).value
                            DSnode8buddy1 = ws.cell(row=cell.row, column=11).value
                            DSnode9buddy1 = ws.cell(row=cell.row, column=12).value
                            DSnode10buddy1 = ws.cell(row=cell.row, column=13).value
                            DSnode11buddy1 = ws.cell(row=cell.row, column=14).value
                            DSnode12buddy1 = ws.cell(row=cell.row, column=15).value
                            DSnode13buddy1 = ws.cell(row=cell.row, column=16).value
                            DSnode14buddy1 = ws.cell(row=cell.row, column=17).value
                            DSnode15buddy1 = ws.cell(row=cell.row, column=18).value
                            DSnode16buddy1 = ws.cell(row=cell.row, column=19).value
                            DSnode17buddy1 = ws.cell(row=cell.row, column=20).value
                            DSnode18buddy1 = ws.cell(row=cell.row, column=21).value
                            DSnode19buddy1 = ws.cell(row=cell.row, column=22).value
                            DSnode20buddy1 = ws.cell(row=cell.row, column=23).value
                            DSnode21buddy1 = ws.cell(row=cell.row, column=24).value
                            DSnode22buddy1 = ws.cell(row=cell.row, column=25).value #up to 22 DS to be split

                            
                            #append values in list and separate which DS a buddy will cover and which ones will go for the other buddy
                            DSscheduler1coverbuddy1.extend((DSnode1buddy1,DSnode2buddy1,DSnode3buddy1,DSnode4buddy1,DSnode5buddy1,DSnode6buddy1,DSnode7buddy1,DSnode8buddy1,DSnode9buddy1,DSnode10buddy1,DSnode11buddy1,DSnode12buddy1,DSnode13buddy1,DSnode14buddy1,DSnode15buddy1,DSnode16buddy1,DSnode17buddy1,DSnode18buddy1,DSnode19buddy1,DSnode20buddy1,DSnode21buddy1,DSnode22buddy1))
                            

                            #getting the country for each DS. Depending on the country it will be displayed with one colour or another       
                            shiftconfiglocation = r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\TaskList\TasklistGeneration\\' + dateStr2 + '\\shift_config.xlsx'
                            wconfig = openpyxl.load_workbook(shiftconfiglocation)
                            wconfigStations= wconfig.worksheets[1]
                            rangeconfig = wconfigStations._cells_by_col(1,1,2,10000)
                            

                            ##############   Displaying stations for the 1st scheduler covering 1st buddy ################
                            #extractng only the first 4 digits of DS node (removing SD in some DS)
                            global DSscheduler1coverbuddycountry1       

                            DSscheduler1coverbuddycountry1 = []
                            for x in DSscheduler1coverbuddy1:
                                if x != None:
                                    DSscheduler1coverbuddycountry1.append(x[0:4])
                            
                            #getting the country for each DS, up to 11 DS to cover
                            #we will create a list with all countries in the same order as the DS are displayed in the list
                            countrylistbuddy1 = []
                            for DSbuddy in DSscheduler1coverbuddycountry1:
                                wconfigStations= wconfig.worksheets[1]
                                rangeconfig = wconfigStations._cells_by_col(1,1,2,10000)
                                try:    
                                    for row in rangeconfig:
                                        for cell in row:
                                            if (cell.value == DSbuddy):
                                                countrylistbuddy1.append(wconfigStations.cell(row=cell.row, column=2).value)
                                except Exception: 
                                    pass
                            
                            #colour formatting and displaying it
                            global count01
                            count01=5
                            for x,y in zip(countrylistbuddy1,DSscheduler1coverbuddy1):
                                print(x)
                                #print(y)
                                if x == "UK":
                                    colour="orange"
                                if x == "FR":
                                    colour="cyan"
                                if x == "DE":
                                    colour="yellow"
                                if x == "ES":
                                    colour="magenta"
                                if x == "IT":
                                    colour="green"
                                if x == "BE":
                                    colour = "SkyBlue4"
                                if x == "NL":
                                    colour = "pale green"
                                if x == "AT":
                                    colour = "light yellow"
                                
                                
                                DSnodebuddy1label = Label(framebuddy1, text = y,padx=20, pady=5, relief=GROOVE, justify=CENTER,width=10, background= colour)
                                DSnodebuddy1label.grid(row=(count01), column=0, columnspan=3, pady=10, padx=10)
                                count01 +=1
                            
                                
            
            ######################################################################################
            #searching for the DS of the second buddy 
            ######################################################################################################################################
                
            #getting the DS for 2nd buddy and appending to a list
            DSscheduler1coverbuddy2 = []
            ws = wb.worksheets[0] #sheet number 1
            range = ws.iter_rows()
            for row in range:
                    for cell in row:
                        if (cell.value == search_strbuddy2):
                            DSnode1buddy2 = ws.cell(row=cell.row, column=4).value #get the value from column 4 and the row where DS has been found
                            DSnode2buddy2 = ws.cell(row=cell.row, column=5).value
                            DSnode3buddy2 = ws.cell(row=cell.row, column=6).value
                            DSnode4buddy2 = ws.cell(row=cell.row, column=7).value
                            DSnode5buddy2 = ws.cell(row=cell.row, column=8).value
                            DSnode6buddy2 = ws.cell(row=cell.row, column=9).value
                            DSnode7buddy2 = ws.cell(row=cell.row, column=10).value
                            DSnode8buddy2 = ws.cell(row=cell.row, column=11).value
                            DSnode9buddy2 = ws.cell(row=cell.row, column=12).value
                            DSnode10buddy2 = ws.cell(row=cell.row, column=13).value
                            DSnode11buddy2 = ws.cell(row=cell.row, column=14).value
                            DSnode12buddy2 = ws.cell(row=cell.row, column=15).value
                            DSnode13buddy2 = ws.cell(row=cell.row, column=16).value
                            DSnode14buddy2 = ws.cell(row=cell.row, column=17).value
                            DSnode15buddy2 = ws.cell(row=cell.row, column=18).value
                            DSnode16buddy2 = ws.cell(row=cell.row, column=19).value
                            DSnode17buddy2 = ws.cell(row=cell.row, column=20).value
                            DSnode18buddy2 = ws.cell(row=cell.row, column=21).value
                            DSnode19buddy2 = ws.cell(row=cell.row, column=22).value
                            DSnode20buddy2 = ws.cell(row=cell.row, column=23).value
                            DSnode21buddy2 = ws.cell(row=cell.row, column=24).value
                            DSnode22buddy2 = ws.cell(row=cell.row, column=25).value #up to 22 DS to be split

                            #append values in list and separate which DS a buddy will cover and which ones will go for the other buddy
                            DSscheduler1coverbuddy2.extend((DSnode1buddy2,DSnode2buddy2,DSnode3buddy2,DSnode4buddy2,DSnode5buddy2,DSnode6buddy2,DSnode7buddy2,DSnode8buddy2,DSnode9buddy2,DSnode10buddy2,DSnode11buddy2,DSnode12buddy2,DSnode13buddy2,DSnode14buddy2,DSnode15buddy2,DSnode16buddy2,DSnode17buddy2,DSnode18buddy2,DSnode19buddy2,DSnode20buddy2,DSnode21buddy2))
                            

                            #getting the country for each DS. Depending on the country it will be displayed with one colour or another       
                            shiftconfiglocation = r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\TaskList\TasklistGeneration\\' + dateStr2 + '\\shift_config.xlsx'
                            wconfig = openpyxl.load_workbook(shiftconfiglocation)
                            wconfigStations= wconfig.worksheets[1]
                            rangeconfig = wconfigStations._cells_by_col(1,1,2,10000)
                            

                            ##############   Displaying stations for the 1st scheduler covering 1st buddy ################
                            #extractng only the first 4 digits of DS node (removing SD in some DS)
                            DSscheduler1coverbuddycountry2 = []
                            for x in DSscheduler1coverbuddy2:
                                if x != None:
                                    DSscheduler1coverbuddycountry2.append(x[0:4])
                            
                            #getting the country for each DS, up to 11 DS to cover
                            #we will create a list with all countries in the same order as the DS are displayed in the list
                            countrylistbuddy2 = []
                            for DSbuddy in DSscheduler1coverbuddycountry2:
                                wconfigStations= wconfig.worksheets[1]
                                rangeconfig = wconfigStations._cells_by_col(1,1,2,10000)
                                try:    
                                    for row in rangeconfig:
                                        for cell in row:
                                            if (cell.value == DSbuddy):
                                                countrylistbuddy2.append(wconfigStations.cell(row=cell.row, column=2).value)
                                except Exception: 
                                    pass
                            
                            #colour formatting and displaying it
                            global count02
                            count02=5
                            for x,y in zip(countrylistbuddy2,DSscheduler1coverbuddy2):
                                print(x)
                                #print(y)
                                if x == "UK":
                                    colour="orange"
                                if x == "FR":
                                    colour="cyan"
                                if x == "DE":
                                    colour="yellow"
                                if x == "ES":
                                    colour="magenta"
                                if x == "IT":
                                    colour="green"
                                if x == "BE":
                                    colour = "SkyBlue4"
                                if x == "NL":
                                    colour = "pale green"
                                if x == "AT":
                                    colour = "light yellow"
                                
                                
                                DSnodebuddy2label1 = Label(framebuddy2, text = y,padx=20, pady=5, relief=GROOVE, justify=CENTER,width=10, background= colour)
                                DSnodebuddy2label1.grid(row=(count02), column=4, columnspan=3, pady=10, padx=10)
                                count02 +=1
                            

            #getting the breaks and adding to table
            
            os.chdir(r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\TaskList\TasklistGeneration\\' + dateStr2 )
            try:
                breakcsv = csv.reader(open("breakinfo.csv","r"))
                for row in breakcsv:
                    if search_strbuddy1 == row[0]:
                        Breaktime = row
                        BreakScheduler1 = Breaktime[1]
            except:
                pass

            try:
                breakcsv = csv.reader(open("breakinfo.csv","r"))
                for row in breakcsv:
                    if search_strbuddy2 == row[0]:
                        Breaktime2 = row
                        BreakScheduler2 = Breaktime2[1]
            except:
                pass
            
            
        
            try: 
                BreakScheduler1label = Label(framebuddy1,  text= BreakScheduler1, justify=CENTER, padx=20, pady=1, fg="red", font=("Helvetica", 10, 'bold'))
                BreakScheduler1label.grid(row=3, column=0, columnspan=3, pady=5, padx=10)
            except:
                pass
            try:
                BreakScheduler2label = Label(framebuddy2,  text= BreakScheduler2, justify=CENTER, padx=20, pady=1,fg="red", font=("Helvetica", 10, 'bold'))
                BreakScheduler2label.grid(row=3, column=4, columnspan=3, pady=5, padx=10)
            except:
                pass
            


            #get the confirmation and adding to table
            
            
            os.chdir(r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\TaskList\TasklistGeneration\\' + dateStr2 )
            try:
                breakcsv = csv.reader(open("breakinfoconfirmation.csv","r"))
                for row in breakcsv:
                    if search_strbuddy1 == row[0]:
                        confirmation = row
                        ConfirmationScheduler1 = confirmation[1]
            except:
                pass

            try:
                breakcsv = csv.reader(open("breakinfoconfirmation.csv","r"))
                for row in breakcsv:
                    if search_strbuddy2 == row[0]:
                        confirmation2 = row
                        ConfirmationScheduler2 = confirmation2[1]
            except:
                pass
            
            

            
            try:
                if ConfirmationScheduler1 != None:
                    ConfirmationScheduler1label = Label(framebuddy1, text= search_strbuddy1 + " confirms covering breaks", foreground= "green", relief=GROOVE, justify=CENTER, padx=20, pady=5)
                    ConfirmationScheduler1label.grid(column = 0, row= count02+2, pady=10, padx=10, columnspan=3)
                    print("confirmed")
            except:
                pass
            try:
                if ConfirmationScheduler2 != None:
                    ConfirmationScheduler2label = Label(framebuddy2, text= search_strbuddy2 + " confirms covering breaks", foreground= "green", relief=GROOVE, justify=CENTER, padx=20, pady=5)
                    ConfirmationScheduler2label.grid(column = 4, row= count01+2, pady=10, padx=10, columnspan=3)
                    print("confirmed")
            except:
                pass
            
            #addding extracover to table
            breakcsv = csv.reader(open("extracover.csv","r"))
            for row in breakcsv:
                if search_strbuddy1 == row[0]:
                    extracoverDS = row
                    extra1cover = extracoverDS[1] 
                    extra2cover = extracoverDS[2] 
                    print(extra1cover)
                    
                    extra1coverlabelinfo = Label(framebuddy1, text= search_strbuddy1 + " has extra cover", relief=GROOVE, justify=CENTER, padx=20, pady=5)
                    extra1coverlabelinfo.grid(column = 0, row= count01+3, pady=10, padx=10, columnspan=3)
                    if extra1cover != "":
                        extra1coverlabel = Label(framebuddy1, text= extra1cover, relief=GROOVE, justify=CENTER, padx=20, pady=5)
                        extra1coverlabel.grid(column = 0, row= count01+4, pady=10, padx=10, columnspan=3) 
                        DSscheduler1coverbuddy1.append(extra1cover) #appendig to list
                    if extra2cover != "":
                        extra2coverlabel = Label(framebuddy1, text= extra2cover, relief=GROOVE, justify=CENTER, padx=20, pady=5)
                        extra2coverlabel.grid(column = 0, row= count01+5, pady=10, padx=10, columnspan=3)
                        DSscheduler1coverbuddy1.append(extra2cover)
                
            
            breakcsv = csv.reader(open("extracover.csv","r"))
            for row in breakcsv:
                if search_strbuddy2 == row[0]:
                    extracoverDS2 = row
                    extra1cover2 = extracoverDS2[1] 
                    extra2cover2 = extracoverDS2[2] 
                    
                    extra1coverlabelinfo = Label(framebuddy2, text= search_strbuddy2 + " has extra cover", relief=GROOVE, justify=CENTER, padx=20, pady=5)
                    extra1coverlabelinfo.grid(column = 4, row= count02+3, pady=10, padx=10, columnspan=3)
                    if extra1cover2 != "":
                        extra1coverlabel2 = Label(framebuddy2, text= extra1cover2, relief=GROOVE, justify=CENTER, padx=20, pady=5,width=10)
                        extra1coverlabel2.grid(column = 4, row= count02+4, pady=10, padx=10, columnspan=3) 
                        DSscheduler1coverbuddy2.append(extra1cover2)
                    if extra2cover2 != "":
                        extra2coverlabel2 = Label(framebuddy2, text= extra2cover2, relief=GROOVE, justify=CENTER, padx=20, pady=5, width=10)
                        extra2coverlabel2.grid(column = 4, row= count02+5, pady=10, padx=10, columnspan=3)
                        DSscheduler1coverbuddy2.append(extra2cover2)
            
            




















        ##########################################################################################################################
        ##########################################################################################################################
        #3 buddy system#
        if Buddy3 != None and Buddy4 == None:
            print("3 Buddy system")
            Buddy1 = Buddy1list[2] #take the 3rd element from the buddylist to get just 1 name
            Buddy2 = Buddy2list[2]
            Buddy3 = Buddy3list[2]
            

            #creating frames for each buddy (3 frames in total)
            framebuddy1 = ttk.LabelFrame(frame.scrollable_frame, padding=5)
            framebuddy1.grid(row=1, column = 0, columnspan=2, \
                padx=5, pady=5, ipadx=5, ipady=5, sticky=N)
            
            framebuddy2 = ttk.LabelFrame(frame.scrollable_frame, padding=5)
            framebuddy2.grid(row=1, column = 3, columnspan=3, \
                padx=5, pady=5, ipadx=5, ipady=5,sticky=N)

            framebuddy3 = ttk.LabelFrame(frame.scrollable_frame, padding=5)
            framebuddy3.grid(row=1, column = 6, columnspan=3, \
                padx=5, pady=5, ipadx=5, ipady=5, sticky=N)


            textlabelbuddy1 = Label(framebuddy1,  text= "Break for " + Buddy1, justify=CENTER, padx=20, pady=5, font=("Helvetica", 12, 'bold'))
            textlabelbuddy1.grid(row=2, column=0, columnspan=2, pady=10, padx=10)
            textlabelbuddy1cover4 = Label(framebuddy1,  text= Buddy2, padx=20, pady=5)
            textlabelbuddy1cover4.grid(row=4, column=0, columnspan=1, pady=10, padx=10)
            textlabelbuddy1cover2 = Label(framebuddy1,  text= Buddy3, padx=20, pady=5)
            textlabelbuddy1cover2.grid(row=4, column=1, columnspan=1, pady=10, padx=10)
            

            textlabelbuddy2 = Label(framebuddy2,  text= "Break for " + Buddy2 , justify=CENTER, padx=20, pady=5, font=("Helvetica", 12, 'bold'))
            textlabelbuddy2.grid(row=2, column=3, columnspan=2, pady=10, padx=10)
            textlabelbuddy2cover1 = Label(framebuddy2,  text= Buddy3, padx=20, pady=5)
            textlabelbuddy2cover1.grid(row=4, column=3, columnspan=1, pady=10, padx=10)
            textlabelbuddy2cover3 = Label(framebuddy2,  text= Buddy1, padx=20, pady=5)
            textlabelbuddy2cover3.grid(row=4, column=4, columnspan=1, pady=10, padx=10)
            

            textlabelbuddy3 = Label(framebuddy3,  text= "Break for " + Buddy3, justify=CENTER, padx=20, pady=5, font=("Helvetica", 12, 'bold'))
            textlabelbuddy3.grid(row=2, column=6, columnspan=3, pady=10, padx=10)
            textlabelbuddy3cover1 = Label(framebuddy3,  text= Buddy1, padx=20, pady=5)
            textlabelbuddy3cover1.grid(row=4, column=6, columnspan=1, pady=10, padx=10)
            textlabelbuddy3cover2 = Label(framebuddy3,  text= Buddy2, padx=20, pady=5)
            textlabelbuddy3cover2.grid(row=4, column=7, columnspan=1, pady=10, padx=10)
            
            ######################################################################################################################################
            #searching for the DS of the first buddy
            search_strbuddy1 = Buddy1
            search_strbuddy2 = Buddy2
            search_strbuddy3 = Buddy3
            
            #getting the DS for 1st buddy and appending to a list
            DSscheduler1coverbuddy1 = []
            DSscheduler2coverbuddy1 = []
            ws = wb.worksheets[0] #sheet number 1
            range = ws.iter_rows()
            for row in range:
                    for cell in row:
                        if (cell.value == search_strbuddy1):
                            DSnode1buddy1 = ws.cell(row=cell.row, column=4).value #get the value from column 4 and the row where DS has been found
                            DSnode2buddy1 = ws.cell(row=cell.row, column=5).value
                            DSnode3buddy1 = ws.cell(row=cell.row, column=6).value
                            DSnode4buddy1 = ws.cell(row=cell.row, column=7).value
                            DSnode5buddy1 = ws.cell(row=cell.row, column=8).value
                            DSnode6buddy1 = ws.cell(row=cell.row, column=9).value
                            DSnode7buddy1 = ws.cell(row=cell.row, column=10).value
                            DSnode8buddy1 = ws.cell(row=cell.row, column=11).value
                            DSnode9buddy1 = ws.cell(row=cell.row, column=12).value
                            DSnode10buddy1 = ws.cell(row=cell.row, column=13).value
                            DSnode11buddy1 = ws.cell(row=cell.row, column=14).value
                            DSnode12buddy1 = ws.cell(row=cell.row, column=15).value
                            DSnode13buddy1 = ws.cell(row=cell.row, column=16).value
                            DSnode14buddy1 = ws.cell(row=cell.row, column=17).value
                            DSnode15buddy1 = ws.cell(row=cell.row, column=18).value
                            DSnode16buddy1 = ws.cell(row=cell.row, column=19).value
                            DSnode17buddy1 = ws.cell(row=cell.row, column=20).value
                            DSnode18buddy1 = ws.cell(row=cell.row, column=21).value
                            DSnode19buddy1 = ws.cell(row=cell.row, column=22).value
                            DSnode20buddy1 = ws.cell(row=cell.row, column=23).value
                            DSnode21buddy1 = ws.cell(row=cell.row, column=24).value
                            DSnode22buddy1 = ws.cell(row=cell.row, column=25).value #up to 22 DS to be split

                            #append values in list and separate which DS a buddy will cover and which ones will go for the other buddy
                            DSscheduler1coverbuddy1.extend((DSnode1buddy1,DSnode3buddy1,DSnode5buddy1,DSnode7buddy1,DSnode9buddy1,DSnode11buddy1,DSnode13buddy1,DSnode15buddy1,DSnode17buddy1,DSnode19buddy1,DSnode21buddy1))
                            DSscheduler2coverbuddy1.extend((DSnode2buddy1,DSnode4buddy1,DSnode6buddy1,DSnode8buddy1,DSnode10buddy1,DSnode12buddy1,DSnode14buddy1,DSnode16buddy1,DSnode18buddy1,DSnode20buddy1,DSnode22buddy1))

                            #getting the country for each DS. Depending on the country it will be displayed with one colour or another       
                            shiftconfiglocation = r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\TaskList\TasklistGeneration\\' + dateStr2 + '\\shift_config.xlsx'
                            wconfig = openpyxl.load_workbook(shiftconfiglocation)
                            wconfigStations= wconfig.worksheets[1]
                            rangeconfig = wconfigStations._cells_by_col(1,1,2,10000)
                            

                            ##############   Displaying stations for the 1st scheduler covering 1st buddy ################
                            #extractng only the first 4 digits of DS node (removing SD in some DS)     

                            DSscheduler1coverbuddycountry1 = []
                            for x in DSscheduler1coverbuddy1:
                                if x != None:
                                    DSscheduler1coverbuddycountry1.append(x[0:4])
                            
                                
                            
                            #getting the country for each DS, up to 11 DS to cover
                            #we will create a list with all countries in the same order as the DS are displayed in the list
                            countrylistbuddy1 = []
                            
                            for DSbuddy in DSscheduler1coverbuddycountry1:
                                wconfigStations= wconfig.worksheets[1]
                                rangeconfig = wconfigStations._cells_by_col(1,1,2,10000)
                                try:    
                                    for row in rangeconfig:
                                        for cell in row:
                                            if (cell.value == DSbuddy):
                                                countrylistbuddy1.append(wconfigStations.cell(row=cell.row, column=2).value)
                                except Exception: 
                                    pass
                            
                            
                            #colour formatting and displaying it
                            global count
                            count=5
                            for x,y in zip(countrylistbuddy1,DSscheduler1coverbuddy1):
                                print(x)
                                #print(y)
                                if x == "UK":
                                    colour="orange"
                                if x == "FR":
                                    colour="cyan"
                                if x == "DE":
                                    colour="yellow"
                                if x == "ES":
                                    colour="magenta"
                                if x == "IT":
                                    colour="green"
                                if x == "BE":
                                    colour = "SkyBlue4"
                                if x == "NL":
                                    colour = "pale green"
                                if x == "AT":
                                    colour = "light yellow"
                                
                                
                                DSnodebuddy1label = Label(framebuddy1, text = y,padx=20, pady=5, relief=GROOVE, width=8, background= colour)
                                DSnodebuddy1label.grid(row=(count), column=0, columnspan=1, pady=10, padx=10)
                                count +=1
                            


                            ##############   Displaying stations for the 1st scheduler covering 1st buddy ################
                            #extractng only the first 4 digits of DS node (removing SD)
                            DSscheduler2coverbuddycountry1 = []
                            for x in DSscheduler2coverbuddy1:
                                if x != None:
                                    DSscheduler2coverbuddycountry1.append(x[0:4])
                            
                            
                            #getting the country for each DS, up to 11 DS to cover
                            #we will create a list with all countries in the same order as the DS are displayed in the list
                            countrylist2buddy1 = []
                            for DSbuddy in DSscheduler2coverbuddycountry1:
                                wconfigStations= wconfig.worksheets[1]
                                rangeconfig = wconfigStations._cells_by_col(1,1,2,10000)
                                try:    
                                    for row in rangeconfig:
                                        for cell in row:
                                            if (cell.value == DSbuddy):
                                                countrylist2buddy1.append(wconfigStations.cell(row=cell.row, column=2).value)
                                                print(countrylist2buddy1)
                                except Exception: 
                                    pass
                            
                            
                            
                            #colour formatting and displaying it
                            global count2
                            count2 = 5
                            for x,y in zip(countrylist2buddy1,DSscheduler2coverbuddy1):
                                #print(x)
                                #print(y)
                                if x == "UK":
                                    colour="orange"
                                if x == "FR":
                                    colour="cyan"
                                if x == "DE":
                                    colour="yellow"
                                if x == "ES":
                                    colour="magenta"
                                if x == "IT":
                                    colour="green"
                                if x == "BE":
                                    colour = "SkyBlue4"
                                if x == "NL":
                                    colour = "pale green"
                                if x == "AT":
                                    colour = "light yellow"
                                
                                
                                DSnodebuddy1label2 = Label(framebuddy1, text = y,padx=20, pady=5, relief=GROOVE, width=8, background= colour)
                                DSnodebuddy1label2.grid(row=(count2), column=1, columnspan=1, pady=10, padx=10)
                                count2 +=1
            
                
            
            
            ######################################################################################
            #searching for the DS of the second buddy 
            ######################################################################################################################################
                
            #getting the DS for 2nd buddy and appending to a list
            DSscheduler1coverbuddy2 = []
            DSscheduler2coverbuddy2 = []
            ws = wb.worksheets[0] #sheet number 1
            range = ws.iter_rows()
            for row in range:
                    for cell in row:
                        if (cell.value == search_strbuddy2):
                            DSnode1buddy2 = ws.cell(row=cell.row, column=4).value #get the value from column 4 and the row where DS has been found
                            DSnode2buddy2 = ws.cell(row=cell.row, column=5).value
                            DSnode3buddy2 = ws.cell(row=cell.row, column=6).value
                            DSnode4buddy2 = ws.cell(row=cell.row, column=7).value
                            DSnode5buddy2 = ws.cell(row=cell.row, column=8).value
                            DSnode6buddy2 = ws.cell(row=cell.row, column=9).value
                            DSnode7buddy2 = ws.cell(row=cell.row, column=10).value
                            DSnode8buddy2 = ws.cell(row=cell.row, column=11).value
                            DSnode9buddy2 = ws.cell(row=cell.row, column=12).value
                            DSnode10buddy2 = ws.cell(row=cell.row, column=13).value
                            DSnode11buddy2 = ws.cell(row=cell.row, column=14).value
                            DSnode12buddy2 = ws.cell(row=cell.row, column=15).value
                            DSnode13buddy2 = ws.cell(row=cell.row, column=16).value
                            DSnode14buddy2 = ws.cell(row=cell.row, column=17).value
                            DSnode15buddy2 = ws.cell(row=cell.row, column=18).value
                            DSnode16buddy2 = ws.cell(row=cell.row, column=19).value
                            DSnode17buddy2 = ws.cell(row=cell.row, column=20).value
                            DSnode18buddy2 = ws.cell(row=cell.row, column=21).value
                            DSnode19buddy2 = ws.cell(row=cell.row, column=22).value
                            DSnode20buddy2 = ws.cell(row=cell.row, column=23).value
                            DSnode21buddy2 = ws.cell(row=cell.row, column=24).value
                            DSnode22buddy2 = ws.cell(row=cell.row, column=25).value #up to 22 DS to be split

                            #append values in list and separate which DS a buddy will cover and which ones will go for the other buddy
                            DSscheduler1coverbuddy2.extend((DSnode1buddy2,DSnode3buddy2,DSnode5buddy2,DSnode7buddy2,DSnode9buddy2,DSnode11buddy2,DSnode13buddy2,DSnode15buddy2,DSnode17buddy2,DSnode19buddy2,DSnode21buddy2))
                            DSscheduler2coverbuddy2.extend((DSnode2buddy2,DSnode4buddy2,DSnode6buddy2,DSnode8buddy2,DSnode10buddy2,DSnode12buddy2,DSnode14buddy2,DSnode16buddy2,DSnode18buddy2,DSnode20buddy2,DSnode22buddy2))

                            #getting the country for each DS. Depending on the country it will be displayed with one colour or another       
                            shiftconfiglocation = r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\TaskList\TasklistGeneration\\' + dateStr2 + '\\shift_config.xlsx'
                            wconfig = openpyxl.load_workbook(shiftconfiglocation)
                            wconfigStations= wconfig.worksheets[1]
                            rangeconfig = wconfigStations._cells_by_col(1,1,2,10000)
                            

                            ##############   Displaying stations for the 1st scheduler covering 1st buddy ################
                            #extractng only the first 4 digits of DS node (removing SD in some DS)
                            DSscheduler1coverbuddycountry2 = []
                            for x in DSscheduler1coverbuddy2:
                                if x != None:
                                    DSscheduler1coverbuddycountry2.append(x[0:4])
                            
                            #getting the country for each DS, up to 11 DS to cover
                            #we will create a list with all countries in the same order as the DS are displayed in the list
                            countrylistbuddy2 = []
                            for DSbuddy in DSscheduler1coverbuddycountry2:
                                wconfigStations= wconfig.worksheets[1]
                                rangeconfig = wconfigStations._cells_by_col(1,1,2,10000)
                                try:    
                                    for row in rangeconfig:
                                        for cell in row:
                                            if (cell.value == DSbuddy):
                                                countrylistbuddy2.append(wconfigStations.cell(row=cell.row, column=2).value)
                                except Exception: 
                                    pass
                            
                            
                            #colour formatting and displaying it
                            global count3
                            count3=5
                            for x,y in zip(countrylistbuddy2,DSscheduler1coverbuddy2):
                                print(x)
                                #print(y)
                                if x == "UK":
                                    colour="orange"
                                if x == "FR":
                                    colour="cyan"
                                if x == "DE":
                                    colour="yellow"
                                if x == "ES":
                                    colour="magenta"
                                if x == "IT":
                                    colour="green"
                                if x == "BE":
                                    colour = "SkyBlue4"
                                if x == "NL":
                                    colour = "pale green"
                                if x == "AT":
                                    colour = "light yellow"
                                
                                
                                DSnodebuddy2label1 = Label(framebuddy2, text = y,padx=20, pady=5, relief=GROOVE, width=8, background= colour)
                                DSnodebuddy2label1.grid(row=(count3), column=3, columnspan=1, pady=10, padx=10)
                                count3 +=1
                            


                            ##############   Displaying stations for the 1st scheduler covering 1st buddy ################
                            #extractng only the first 4 digits of DS node (removing SD)
                            DSscheduler2coverbuddycountry2 = []
                            for x in DSscheduler2coverbuddy2:
                                if x != None:
                                    DSscheduler2coverbuddycountry2.append(x[0:4])
                            
                            #getting the country for each DS, up to 11 DS to cover
                            #we will create a list with all countries in the same order as the DS are displayed in the list
                            countrylist2buddy2 = []
                            for DSforbuddy in DSscheduler2coverbuddycountry2:
                                wconfigStations= wconfig.worksheets[1]
                                rangeconfig = wconfigStations._cells_by_col(1,1,2,10000)
                                try:    
                                    for row in rangeconfig:
                                        for cell in row:
                                            if (cell.value == DSforbuddy):
                                                countrylist2buddy2.append(wconfigStations.cell(row=cell.row, column=2).value)
                                except Exception: 
                                    pass
                            
                            
                            
                            #colour formatting and displaying it
                            global count4
                            count4 = 5
                            for x,y in zip(countrylist2buddy2,DSscheduler2coverbuddy2):
                                #print(x)
                                #print(y)
                                if x == "UK":
                                    colour="orange"
                                if x == "FR":
                                    colour="cyan"
                                if x == "DE":
                                    colour="yellow"
                                if x == "ES":
                                    colour="magenta"
                                if x == "IT":
                                    colour="green"
                                if x == "BE":
                                    colour = "SkyBlue4"
                                if x == "NL":
                                    colour = "pale green"
                                if x == "AT":
                                    colour = "light yellow"
                                
                                
                                DSnodebuddy2label2 = Label(framebuddy2, text = y,padx=20, pady=5, relief=GROOVE, width=8, background= colour)
                                DSnodebuddy2label2.grid(row=(count4), column=4, columnspan=1, pady=10, padx=10)
                                count4 +=1

                        
        ######################################################################################
            #searching for the DS of the third buddy 
        ######################################################################################################################################
                
            #getting the DS for 2nd buddy and appending to a list
            DSscheduler1coverbuddy3 = []
            DSscheduler2coverbuddy3 = []
            ws = wb.worksheets[0] #sheet number 1
            range = ws.iter_rows()
            for row in range:
                    for cell in row:
                        if (cell.value == search_strbuddy3):
                            DSnode1buddy3 = ws.cell(row=cell.row, column=4).value #get the value from column 4 and the row where DS has been found
                            DSnode2buddy3 = ws.cell(row=cell.row, column=5).value
                            DSnode3buddy3 = ws.cell(row=cell.row, column=6).value
                            DSnode4buddy3 = ws.cell(row=cell.row, column=7).value
                            DSnode5buddy3 = ws.cell(row=cell.row, column=8).value
                            DSnode6buddy3 = ws.cell(row=cell.row, column=9).value
                            DSnode7buddy3 = ws.cell(row=cell.row, column=10).value
                            DSnode8buddy3 = ws.cell(row=cell.row, column=11).value
                            DSnode9buddy3 = ws.cell(row=cell.row, column=12).value
                            DSnode10buddy3 = ws.cell(row=cell.row, column=13).value
                            DSnode11buddy3 = ws.cell(row=cell.row, column=14).value
                            DSnode12buddy3 = ws.cell(row=cell.row, column=15).value
                            DSnode13buddy3 = ws.cell(row=cell.row, column=16).value
                            DSnode14buddy3 = ws.cell(row=cell.row, column=17).value
                            DSnode15buddy3 = ws.cell(row=cell.row, column=18).value
                            DSnode16buddy3 = ws.cell(row=cell.row, column=19).value
                            DSnode17buddy3 = ws.cell(row=cell.row, column=20).value
                            DSnode18buddy3 = ws.cell(row=cell.row, column=21).value
                            DSnode19buddy3 = ws.cell(row=cell.row, column=22).value
                            DSnode20buddy3 = ws.cell(row=cell.row, column=23).value
                            DSnode21buddy3 = ws.cell(row=cell.row, column=24).value
                            DSnode22buddy3 = ws.cell(row=cell.row, column=25).value #up to 22 DS to be split

                            #append values in list and separate which DS a buddy will cover and which ones will go for the other buddy
                            DSscheduler1coverbuddy3.extend((DSnode1buddy3,DSnode3buddy3,DSnode5buddy3,DSnode7buddy3,DSnode9buddy3,DSnode11buddy3,DSnode13buddy3,DSnode15buddy3,DSnode17buddy3,DSnode19buddy3,DSnode21buddy3))
                            DSscheduler2coverbuddy3.extend((DSnode2buddy3,DSnode4buddy3,DSnode6buddy3,DSnode8buddy3,DSnode10buddy3,DSnode12buddy3,DSnode14buddy3,DSnode16buddy3,DSnode18buddy3,DSnode20buddy3,DSnode22buddy3))

                            #getting the country for each DS. Depending on the country it will be displayed with one colour or another       
                            shiftconfiglocation = r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\TaskList\TasklistGeneration\\' + dateStr2 + '\\shift_config.xlsx'
                            wconfig = openpyxl.load_workbook(shiftconfiglocation)
                            wconfigStations= wconfig.worksheets[1]
                            rangeconfig = wconfigStations._cells_by_col(1,1,2,10000)
                            

                            ##############   Displaying stations for the 1st scheduler covering 3rd buddy ################
                            #extractng only the first 4 digits of DS node (removing SD in some DS)
                            DSscheduler1coverbuddycountry3 = []
                            for x in DSscheduler1coverbuddy3:
                                if x != None:
                                    DSscheduler1coverbuddycountry3.append(x[0:4])
                            
                            #getting the country for each DS, up to 11 DS to cover
                            #we will create a list with all countries in the same order as the DS are displayed in the list
                            countrylistbuddy3 = []
                            for DSforbuddy in DSscheduler1coverbuddycountry3:
                                wconfigStations= wconfig.worksheets[1]
                                rangeconfig = wconfigStations._cells_by_col(1,1,2,10000)
                                try:    
                                    for row in rangeconfig:
                                        for cell in row:
                                            if (cell.value == DSforbuddy):
                                                countrylistbuddy3.append(wconfigStations.cell(row=cell.row, column=2).value)
                                except Exception: 
                                    pass
                            


                            #colour formatting and displaying it
                            global count5
                            count5=5
                            for x,y in zip(countrylistbuddy3,DSscheduler1coverbuddy3):
                                print(x)
                                #print(y)
                                if x == "UK":
                                    colour="orange"
                                if x == "FR":
                                    colour="cyan"
                                if x == "DE":
                                    colour="yellow"
                                if x == "ES":
                                    colour="magenta"
                                if x == "IT":
                                    colour="green"
                                if x == "BE":
                                    colour = "SkyBlue4"
                                if x == "NL":
                                    colour = "pale green"
                                if x == "AT":
                                    colour = "light yellow"
                                
                                
                                DSnodebuddy3label1 = Label(framebuddy3, text = y,padx=20, pady=5, relief=GROOVE, width=8, background= colour)
                                DSnodebuddy3label1.grid(row=(count5), column=6, columnspan=1, pady=10, padx=10)
                                count5 +=1
                            


                            ##############   Displaying stations for the 2nd scheduler covering 3rd buddy ################
                            #extractng only the first 4 digits of DS node (removing SD)
                            DSscheduler2coverbuddycountry3 = []
                            for x in DSscheduler2coverbuddy3:
                                if x != None:
                                    DSscheduler2coverbuddycountry3.append(x[0:4])
                            
                            #getting the country for each DS, up to 11 DS to cover
                            #we will create a list with all countries in the same order as the DS are displayed in the list
                            countrylist2buddy3 = []
                            for DSforbuddy in DSscheduler2coverbuddycountry3:
                                wconfigStations= wconfig.worksheets[1]
                                rangeconfig = wconfigStations._cells_by_col(1,1,2,10000)
                                try:    
                                    for row in rangeconfig:
                                        for cell in row:
                                            if (cell.value == DSforbuddy):
                                                countrylist2buddy3.append(wconfigStations.cell(row=cell.row, column=2).value)
                                except Exception: 
                                    pass
                            
                            #colour formatting and displaying it
                            global count6
                            count6 = 5
                            for x,y in zip(countrylist2buddy3,DSscheduler2coverbuddy3):
                                #print(x)
                                #print(y)
                                if x == "UK":
                                    colour="orange"
                                if x == "FR":
                                    colour="cyan"
                                if x == "DE":
                                    colour="yellow"
                                if x == "ES":
                                    colour="magenta"
                                if x == "IT":
                                    colour="green"
                                if x == "BE":
                                    colour = "SkyBlue4"
                                if x == "NL":
                                    colour = "pale green"
                                if x == "AT":
                                    colour = "light yellow"
                                
                                
                                DSnodebuddy3label2 = Label(framebuddy3, text = y,padx=20, pady=5, relief=GROOVE, width=8, background= colour)
                                DSnodebuddy3label2.grid(row=(count6), column=7, columnspan=1, pady=10, padx=10)
                                count6 +=1
                                                
            

            #getting the extracovers and adding to table
            breakcsv = csv.reader(open("extracover.csv","r"))
            for row in breakcsv:
                if search_strbuddy1 == row[0]:
                    extracoverDS = row
                    extra1cover = extracoverDS[1] 
                    extra2cover = extracoverDS[2] 
                    print(extra1cover)
                    
                    extra1coverlabelinfo = Label(framebuddy1, text= search_strbuddy1 + " has extra cover", relief=GROOVE, justify=CENTER, padx=20, pady=5)
                    extra1coverlabelinfo.grid(column = 0, row= count+3, pady=10, padx=10, columnspan=2)
                    if extra1cover != "":
                        extra1coverlabel = Label(framebuddy1, text= extra1cover, relief=GROOVE, justify=CENTER, padx=20, pady=5)
                        extra1coverlabel.grid(column = 1, row= count+4, pady=10, padx=10, columnspan=1)
                        DSscheduler2coverbuddycountry1.append(extra1cover) 
                    if extra2cover != "":
                        extra2coverlabel = Label(framebuddy1, text= extra2cover, relief=GROOVE, justify=CENTER, padx=20, pady=5)
                        extra2coverlabel.grid(column = 0, row= count+4, pady=10, padx=10, columnspan=1)
                        DSscheduler1coverbuddycountry1.append(extra2cover) 
                
            
            breakcsv = csv.reader(open("extracover.csv","r"))
            for row in breakcsv:
                if search_strbuddy2 == row[0]:
                    extracoverDS2 = row
                    extra1cover2 = extracoverDS2[1] 
                    extra2cover2 = extracoverDS2[2] 
                    
                    extra1coverlabelinfo = Label(framebuddy2, text= search_strbuddy2 + " has extra cover", relief=GROOVE, justify=CENTER, padx=20, pady=5)
                    extra1coverlabelinfo.grid(column = 3, row= count3+3, pady=10, padx=10, columnspan=2)
                    if extra1cover2 != "":
                        extra1coverlabel2 = Label(framebuddy2, text= extra1cover2, relief=GROOVE, justify=CENTER, padx=20, pady=5)
                        extra1coverlabel2.grid(column = 4, row= count3+4, pady=10, padx=10, columnspan=1) 
                        DSscheduler2coverbuddycountry2.append(extra1cover2)
                    if extra2cover2 != "":
                        extra2coverlabel2 = Label(framebuddy2, text= extra2cover2, relief=GROOVE, justify=CENTER, padx=20, pady=5)
                        extra2coverlabel2.grid(column = 3, row= count3+4, pady=10, padx=10, columnspan=1)
                        DSscheduler1coverbuddycountry2.append(extra2cover2)
                
                    
            
            breakcsv = csv.reader(open("extracover.csv","r"))
            for row in breakcsv:
                if search_strbuddy3 == row[0]:
                    extracoverDS3 = row
                    extra1cover3 = extracoverDS3[1] 
                    extra2cover3 = extracoverDS3[2]
                    extra1coverlabelinfo = Label(framebuddy3, text= search_strbuddy3 + " has extra cover", relief=GROOVE, justify=CENTER, padx=20, pady=5)
                    extra1coverlabelinfo.grid(column = 6, row= count5+3, pady=10, padx=10, columnspan=2)
                    if extra1cover3 != "":
                        extra1coverlabel3 = Label(framebuddy3, text= extra1cover3, relief=GROOVE, justify=CENTER, padx=20, pady=5)
                        extra1coverlabel3.grid(column = 7, row= count5+4, pady=10, padx=10, columnspan=1) 
                        DSscheduler2coverbuddycountry3.append(extra1cover3)
                    if extra2cover3 != "":
                        extra2coverlabel3 = Label(framebuddy3, text= extra2cover3, relief=GROOVE, justify=CENTER, padx=20, pady=5)
                        extra2coverlabel3.grid(column = 6, row= count5+4, pady=10, padx=10, columnspan=1)
                        DSscheduler1coverbuddycountry3.append(extra2cover3)
                    
            
            #getting the breaks and adding to table
            
            os.chdir(r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\TaskList\TasklistGeneration\\' + dateStr2 )
            try:
                breakcsv = csv.reader(open("breakinfo.csv","r"))
                for row in breakcsv:
                    if search_strbuddy1 == row[0]:
                        Breaktime = row
                        BreakScheduler1 = Breaktime[1]
            except:
                pass

            try:
                breakcsv = csv.reader(open("breakinfo.csv","r"))
                for row in breakcsv:
                    if search_strbuddy2 == row[0]:
                        Breaktime2 = row
                        BreakScheduler2 = Breaktime2[1]
            except:
                pass
            try:
                breakcsv = csv.reader(open("breakinfo.csv","r"))
                for row in breakcsv:
                    if search_strbuddy3 == row[0]:
                        Breaktime3 = row
                        BreakScheduler3 = Breaktime3[1]
            except:
                pass
            
        
            try: 
                BreakScheduler1label = Label(framebuddy1,  text= BreakScheduler1, justify=CENTER, padx=20, pady=1, fg="red", font=("Helvetica", 10, 'bold'))
                BreakScheduler1label.grid(row=3, column=0, columnspan=2, pady=5, padx=10)
            except:
                pass
            try:
                BreakScheduler2label = Label(framebuddy2,  text= BreakScheduler2, justify=CENTER, padx=20, pady=1,fg="red", font=("Helvetica", 10, 'bold'))
                BreakScheduler2label.grid(row=3, column=3, columnspan=2, pady=5, padx=10)
            except:
                pass
            try:
                BreakScheduler3label = Label(framebuddy3,  text= BreakScheduler3, justify=CENTER, padx=20, pady=1,fg="red", font=("Helvetica", 10, 'bold'))
                BreakScheduler3label.grid(row=3, column=6, columnspan=2, pady=5, padx=10)
            except:
                pass
        

            #get the confirmation and adding to table
            
            
            os.chdir(r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\TaskList\TasklistGeneration\\' + dateStr2 )
            try:
                breakcsv = csv.reader(open("breakinfoconfirmation.csv","r"))
                for row in breakcsv:
                    if search_strbuddy1 == row[0]:
                        confirmation = row
                        ConfirmationScheduler1 = confirmation[1]
            except:
                pass

            try:
                breakcsv = csv.reader(open("breakinfoconfirmation.csv","r"))
                for row in breakcsv:
                    if search_strbuddy2 == row[0]:
                        confirmation2 = row
                        ConfirmationScheduler2 = confirmation2[1]
            except:
                pass
            try:
                breakcsv = csv.reader(open("breakinfoconfirmation.csv","r"))
                for row in breakcsv:
                    if search_strbuddy3 == row[0]:
                        confirmation3 = row
                        ConfirmationScheduler3 = confirmation3[1]
            except:
                pass
            

            
            try:
                if ConfirmationScheduler1 != None:
                    ConfirmationScheduler1label = Label(framebuddy1, text= search_strbuddy1 + " confirms covering breaks", foreground= "green", relief=GROOVE, justify=CENTER, padx=20, pady=5)
                    ConfirmationScheduler1label.grid(column = 0, row= count+2, pady=10, padx=10, columnspan=2)
                    print("confirmed")
            except:
                pass
            try:
                if ConfirmationScheduler2 != None:
                    ConfirmationScheduler2label = Label(framebuddy2, text= search_strbuddy2 + " confirms covering breaks", foreground= "green", relief=GROOVE, justify=CENTER, padx=20, pady=5)
                    ConfirmationScheduler2label.grid(column = 3, row= count+2, pady=10, padx=10, columnspan=2)
                    print("confirmed")
            except:
                pass
            try:
                if ConfirmationScheduler3 != None:
                    ConfirmationScheduler3label = Label(framebuddy3, text= search_strbuddy3 + " confirms covering breaks", foreground= "green", relief=GROOVE, justify=CENTER, padx=20, pady=5)
                    ConfirmationScheduler3label.grid(column = 6, row= count+2, pady=10, padx=10, columnspan=2)
                    print("confirmed")
            except:
                pass
        

        














        ##########################################################################################################################
        ##########################################################################################################################
        #4 buddy system#
        if Buddy4 != None:
            print("4 Buddy System")
            Buddy1 = Buddy1list[3] #take the 4th element from the buddylist to get just 1 name
            Buddy2 = Buddy2list[3]
            Buddy3 = Buddy3list[3]
            Buddy4 = Buddy4list[3]
            
        

            #creating frames for each buddy (3 frames in total)
            framebuddy1 = ttk.LabelFrame(frame.scrollable_frame, padding=5)
            framebuddy1.grid(row=1, column = 0, columnspan=3, \
                padx=5, pady=5, ipadx=5, ipady=5, sticky=N)
            
            framebuddy2 = ttk.LabelFrame(frame.scrollable_frame, padding=5)
            framebuddy2.grid(row=1, column = 3, columnspan=3, \
                padx=5, pady=5, ipadx=5, ipady=5,sticky=N)

            framebuddy3 = ttk.LabelFrame(frame.scrollable_frame, padding=5)
            framebuddy3.grid(row=1, column = 6, columnspan=3, \
                padx=5, pady=5, ipadx=5, ipady=5, sticky=N)
            
            framebuddy4 = ttk.LabelFrame(frame.scrollable_frame, padding=5)
            framebuddy4.grid(row=1, column = 9, columnspan=3, \
                padx=5, pady=5, ipadx=5, ipady=5, sticky=N)


            textlabelbuddy1 = Label(framebuddy1,  text= "Break for " + str(Buddy1), justify=CENTER, padx=20, pady=5, font=("Helvetica", 12, 'bold'))
            textlabelbuddy1.grid(row=2, column=0, columnspan=3, pady=10, padx=10)
            textlabelbuddy1cover2 = Label(framebuddy1,  text= Buddy2, padx=20, pady=5)
            textlabelbuddy1cover2.grid(row=4, column=0, columnspan=1, pady=10, padx=10)
            textlabelbuddy1cover3 = Label(framebuddy1,  text= Buddy3, padx=20, pady=5)
            textlabelbuddy1cover3.grid(row=4, column=1, columnspan=1, pady=10, padx=10)
            textlabelbuddy1cover4 = Label(framebuddy1,  text= Buddy4, padx=20, pady=5)
            textlabelbuddy1cover4.grid(row=4, column=2, columnspan=1, pady=10, padx=10)
            

            textlabelbuddy2 = Label(framebuddy2,  text= "Break for " + str(Buddy2) , justify=CENTER, padx=20, pady=5, font=("Helvetica", 12, 'bold'))
            textlabelbuddy2.grid(row=2, column=3, columnspan=3, pady=10, padx=10)
            textlabelbuddy2cover1 = Label(framebuddy2,  text= Buddy3, padx=20, pady=5)
            textlabelbuddy2cover1.grid(row=4, column=3, columnspan=1, pady=10, padx=10)
            textlabelbuddy2cover3 = Label(framebuddy2,  text= Buddy4, padx=20, pady=5)
            textlabelbuddy2cover3.grid(row=4, column=4, columnspan=1, pady=10, padx=10)
            textlabelbuddy2cover4 = Label(framebuddy2,  text= Buddy1, padx=20, pady=5)
            textlabelbuddy2cover4.grid(row=4, column=5, columnspan=1, pady=10, padx=10)
            

            textlabelbuddy3 = Label(framebuddy3,  text= "Break for " + str(Buddy3), justify=CENTER, padx=20, pady=5, font=("Helvetica", 12, 'bold'))
            textlabelbuddy3.grid(row=2, column=6, columnspan=3, pady=10, padx=10)
            textlabelbuddy3cover1 = Label(framebuddy3,  text= Buddy4, padx=20, pady=5)
            textlabelbuddy3cover1.grid(row=4, column=6, columnspan=1, pady=10, padx=10)
            textlabelbuddy3cover2 = Label(framebuddy3,  text= Buddy1, padx=20, pady=5)
            textlabelbuddy3cover2.grid(row=4, column=7, columnspan=1, pady=10, padx=10)
            textlabelbuddy3cover4 = Label(framebuddy3,  text= Buddy2, padx=20, pady=5)
            textlabelbuddy3cover4.grid(row=4, column=8, columnspan=1, pady=10, padx=10)

            textlabelbuddy4 = Label(framebuddy4,  text= "Break " + str(Buddy4) , justify=LEFT, padx=20, pady=5, font=("Helvetica", 12, 'bold'))
            textlabelbuddy4.grid(row=2, column=9, columnspan=3, pady=10, padx=10)
            textlabelbuddy4cover1 = Label(framebuddy4,  text= Buddy1 , justify=LEFT, padx=20, pady=5)
            textlabelbuddy4cover1.grid(row=4, column=9, columnspan=1, pady=10, padx=10)
            textlabelbuddy4cover2 = Label(framebuddy4,  text= Buddy2 , justify=LEFT, padx=20, pady=5)
            textlabelbuddy4cover2.grid(row=4, column=10, columnspan=1, pady=10, padx=10)
            textlabelbuddy4cover3 = Label(framebuddy4,  text= Buddy3 , justify=LEFT, padx=20, pady=5)
            textlabelbuddy4cover3.grid(row=4, column=11, columnspan=1, pady=10, padx=10)
            
            ######################################################################################################################################
            #searching for the DS of the first buddy
            search_strbuddy1 = Buddy1
            search_strbuddy2 = Buddy2
            search_strbuddy3 = Buddy3
            search_strbuddy4 = Buddy4
            
            #getting the DS for 1st buddy and appending to a list
            DSscheduler1coverbuddy1 = []
            DSscheduler2coverbuddy1 = []
            DSscheduler3coverbuddy1 = []
            ws = wb.worksheets[0] #sheet number 1
            range = ws.iter_rows()
            for row in range:
                    for cell in row:
                        if (cell.value == search_strbuddy1):
                            DSnode1buddy1 = ws.cell(row=cell.row, column=4).value #get the value from column 4 and the row where DS has been found
                            DSnode2buddy1 = ws.cell(row=cell.row, column=5).value
                            DSnode3buddy1 = ws.cell(row=cell.row, column=6).value
                            DSnode4buddy1 = ws.cell(row=cell.row, column=7).value
                            DSnode5buddy1 = ws.cell(row=cell.row, column=8).value
                            DSnode6buddy1 = ws.cell(row=cell.row, column=9).value
                            DSnode7buddy1 = ws.cell(row=cell.row, column=10).value
                            DSnode8buddy1 = ws.cell(row=cell.row, column=11).value
                            DSnode9buddy1 = ws.cell(row=cell.row, column=12).value
                            DSnode10buddy1 = ws.cell(row=cell.row, column=13).value
                            DSnode11buddy1 = ws.cell(row=cell.row, column=14).value
                            DSnode12buddy1 = ws.cell(row=cell.row, column=15).value
                            DSnode13buddy1 = ws.cell(row=cell.row, column=16).value
                            DSnode14buddy1 = ws.cell(row=cell.row, column=17).value
                            DSnode15buddy1 = ws.cell(row=cell.row, column=18).value
                            DSnode16buddy1 = ws.cell(row=cell.row, column=19).value
                            DSnode17buddy1 = ws.cell(row=cell.row, column=20).value
                            DSnode18buddy1 = ws.cell(row=cell.row, column=21).value
                            DSnode19buddy1 = ws.cell(row=cell.row, column=22).value
                            DSnode20buddy1 = ws.cell(row=cell.row, column=23).value
                            DSnode21buddy1 = ws.cell(row=cell.row, column=24).value
                            DSnode22buddy1 = ws.cell(row=cell.row, column=25).value #up to 22 DS to be split

                            #append values in list and separate which DS a buddy will cover and which ones will go for the other buddy
                            DSscheduler1coverbuddy1.extend((DSnode1buddy1,DSnode4buddy1,DSnode7buddy1,DSnode10buddy1,DSnode13buddy1,DSnode16buddy1,DSnode19buddy1,DSnode22buddy1))
                            DSscheduler2coverbuddy1.extend((DSnode2buddy1,DSnode5buddy1,DSnode8buddy1,DSnode11buddy1,DSnode14buddy1,DSnode17buddy1,DSnode20buddy1))
                            DSscheduler3coverbuddy1.extend((DSnode3buddy1,DSnode6buddy1,DSnode9buddy1,DSnode12buddy1,DSnode15buddy1,DSnode18buddy1,DSnode21buddy1))
                            
                            #getting the country for each DS. Depending on the country it will be displayed with one colour or another       
                            shiftconfiglocation = r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\TaskList\TasklistGeneration\\' + dateStr2 + '\\shift_config.xlsx'
                            wconfig = openpyxl.load_workbook(shiftconfiglocation)
                            wconfigStations= wconfig.worksheets[1]
                            rangeconfig = wconfigStations._cells_by_col(1,1,2,10000)
                            

                            ##############   Displaying stations for the 1st scheduler covering 1st buddy ################
                            #extractng only the first 4 digits of DS node (removing SD in some DS)     

                            DSscheduler1coverbuddycountry1 = []
                            for x in DSscheduler1coverbuddy1:
                                if x != None:
                                    DSscheduler1coverbuddycountry1.append(x[0:4])
                            
                            #getting the country for each DS, up to 11 DS to cover
                            #we will create a list with all countries in the same order as the DS are displayed in the list
                            countrylistbuddy1 = []
                            for DSforbuddy in DSscheduler1coverbuddycountry1:
                                wconfigStations= wconfig.worksheets[1]
                                rangeconfig = wconfigStations._cells_by_col(1,1,2,10000)
                                try:    
                                    for row in rangeconfig:
                                        for cell in row:
                                            if (cell.value == DSforbuddy):
                                                countrylistbuddy1.append(wconfigStations.cell(row=cell.row, column=2).value)
                                except Exception: 
                                    pass
                            
                            
                            #colour formatting and displaying it
                            global count001
                            count001=5
                            for x,y in zip(countrylistbuddy1,DSscheduler1coverbuddy1):
                                print(x)
                                #print(y)
                                if x == "UK":
                                    colour="orange"
                                if x == "FR":
                                    colour="cyan"
                                if x == "DE":
                                    colour="yellow"
                                if x == "ES":
                                    colour="magenta"
                                if x == "IT":
                                    colour="green"
                                if x == "BE":
                                    colour = "SkyBlue4"
                                if x == "NL":
                                    colour = "pale green"
                                if x == "AT":
                                    colour = "light yellow"
                                
                                
                                DSnodebuddy1label = Label(framebuddy1, text = y,padx=20, pady=5, relief=GROOVE, width=8, background= colour)
                                DSnodebuddy1label.grid(row=(count001), column=0, columnspan=1, pady=10, padx=10)
                                count001 +=1
                            


                            ##############   Displaying stations for the 1st scheduler covering 1st buddy ################
                            #extractng only the first 4 digits of DS node (removing SD)
                            DSscheduler2coverbuddycountry1 = []
                            for x in DSscheduler2coverbuddy1:
                                if x != None:
                                    DSscheduler2coverbuddycountry1.append(x[0:4])
                            
                            #getting the country for each DS, up to 11 DS to cover
                            #we will create a list with all countries in the same order as the DS are displayed in the list
                            countrylist2buddy1 = []
                            for DSforbuddy in DSscheduler2coverbuddycountry1:
                                wconfigStations= wconfig.worksheets[1]
                                rangeconfig = wconfigStations._cells_by_col(1,1,2,10000)
                                try:    
                                    for row in rangeconfig:
                                        for cell in row:
                                            if (cell.value == DSforbuddy):
                                                countrylist2buddy1.append(wconfigStations.cell(row=cell.row, column=2).value)
                                except Exception: 
                                    pass
                            

                            #colour formatting and displaying it
                            global count002
                            count002 = 5
                            for x,y in zip(countrylist2buddy1,DSscheduler2coverbuddy1):
                                #print(x)
                                #print(y)
                                if x == "UK":
                                    colour="orange"
                                if x == "FR":
                                    colour="cyan"
                                if x == "DE":
                                    colour="yellow"
                                if x == "ES":
                                    colour="magenta"
                                if x == "IT":
                                    colour="green"
                                if x == "BE":
                                    colour = "SkyBlue4"
                                if x == "NL":
                                    colour = "pale green"
                                if x == "AT":
                                    colour = "light yellow"
                                
                                
                                DSnodebuddy1label = Label(framebuddy1, text = y,padx=20, pady=5, relief=GROOVE, width=8, background= colour)
                                DSnodebuddy1label.grid(row=(count002), column=1, columnspan=1, pady=10, padx=10)
                                count002 +=1
                            ##############   Displaying stations for the 3rd scheduler covering 1st buddy ################
                            #extractng only the first 4 digits of DS node (removing SD)
                            DSscheduler3coverbuddycountry1 = []
                            for x in DSscheduler3coverbuddy1:
                                if x != None:
                                    DSscheduler3coverbuddycountry1.append(x[0:4])
                            
                            #getting the country for each DS, up to 11 DS to cover
                            #we will create a list with all countries in the same order as the DS are displayed in the list
                            countrylist3buddy1 = []
                            for DSforbuddy in DSscheduler3coverbuddycountry1:
                                wconfigStations= wconfig.worksheets[1]
                                rangeconfig = wconfigStations._cells_by_col(1,1,2,10000)
                                try:    
                                    for row in rangeconfig:
                                        for cell in row:
                                            if (cell.value == DSforbuddy):
                                                countrylist3buddy1.append(wconfigStations.cell(row=cell.row, column=2).value)
                                except Exception: 
                                    pass
                            
                            
                            
                            #colour formatting and displaying it
                            global count003
                            count003 = 5
                            for x,y in zip(countrylist3buddy1,DSscheduler3coverbuddy1):
                                #print(x)
                                #print(y)
                                if x == "UK":
                                    colour="orange"
                                if x == "FR":
                                    colour="cyan"
                                if x == "DE":
                                    colour="yellow"
                                if x == "ES":
                                    colour="magenta"
                                if x == "IT":
                                    colour="green"
                                if x == "BE":
                                    colour = "SkyBlue4"
                                if x == "NL":
                                    colour = "pale green"
                                if x == "AT":
                                    colour = "light yellow"
                                
                                
                                DSnodebuddy1label = Label(framebuddy1, text = y,padx=20, pady=5, relief=GROOVE, width=8, background= colour)
                                DSnodebuddy1label.grid(row=(count003), column=2, columnspan=1, pady=10, padx=10)
                                count003 +=1
            
                    
            
            
            ######################################################################################
            #searching for the DS of the second buddy 
            ######################################################################################################################################
                
            #getting the DS for 2nd buddy and appending to a list
            DSscheduler1coverbuddy2 = []
            DSscheduler2coverbuddy2 = []
            DSscheduler3coverbuddy2 = []
            ws = wb.worksheets[0] #sheet number 1
            range = ws.iter_rows()
            for row in range:
                    for cell in row:
                        if (cell.value == search_strbuddy2):
                            DSnode1buddy2 = ws.cell(row=cell.row, column=4).value #get the value from column 4 and the row where DS has been found
                            DSnode2buddy2 = ws.cell(row=cell.row, column=5).value
                            DSnode3buddy2 = ws.cell(row=cell.row, column=6).value
                            DSnode4buddy2 = ws.cell(row=cell.row, column=7).value
                            DSnode5buddy2 = ws.cell(row=cell.row, column=8).value
                            DSnode6buddy2 = ws.cell(row=cell.row, column=9).value
                            DSnode7buddy2 = ws.cell(row=cell.row, column=10).value
                            DSnode8buddy2 = ws.cell(row=cell.row, column=11).value
                            DSnode9buddy2 = ws.cell(row=cell.row, column=12).value
                            DSnode10buddy2 = ws.cell(row=cell.row, column=13).value
                            DSnode11buddy2 = ws.cell(row=cell.row, column=14).value
                            DSnode12buddy2 = ws.cell(row=cell.row, column=15).value
                            DSnode13buddy2 = ws.cell(row=cell.row, column=16).value
                            DSnode14buddy2 = ws.cell(row=cell.row, column=17).value
                            DSnode15buddy2 = ws.cell(row=cell.row, column=18).value
                            DSnode16buddy2 = ws.cell(row=cell.row, column=19).value
                            DSnode17buddy2 = ws.cell(row=cell.row, column=20).value
                            DSnode18buddy2 = ws.cell(row=cell.row, column=21).value
                            DSnode19buddy2 = ws.cell(row=cell.row, column=22).value
                            DSnode20buddy2 = ws.cell(row=cell.row, column=23).value
                            DSnode21buddy2 = ws.cell(row=cell.row, column=24).value
                            DSnode22buddy2 = ws.cell(row=cell.row, column=25).value #up to 22 DS to be split

                            #append values in list and separate which DS a buddy will cover and which ones will go for the other buddy
                            DSscheduler1coverbuddy2.extend((DSnode1buddy2,DSnode4buddy2,DSnode7buddy2,DSnode10buddy2,DSnode13buddy2,DSnode16buddy2,DSnode19buddy2,DSnode22buddy2))
                            DSscheduler2coverbuddy2.extend((DSnode2buddy2,DSnode5buddy2,DSnode8buddy2,DSnode11buddy2,DSnode14buddy2,DSnode17buddy2,DSnode20buddy2))
                            DSscheduler3coverbuddy2.extend((DSnode3buddy2,DSnode6buddy2,DSnode9buddy2,DSnode12buddy2,DSnode15buddy2,DSnode18buddy2,DSnode21buddy2))
                            
                            #getting the country for each DS. Depending on the country it will be displayed with one colour or another       
                            shiftconfiglocation = r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\TaskList\TasklistGeneration\\' + dateStr2 + '\\shift_config.xlsx'
                            wconfig = openpyxl.load_workbook(shiftconfiglocation)
                            wconfigStations= wconfig.worksheets[1]
                            rangeconfig = wconfigStations._cells_by_col(1,1,2,10000)
                            

                            ##############   Displaying stations for the 1st scheduler covering 2nd buddy ################
                            #extractng only the first 4 digits of DS node (removing SD in some DS)
                            DSscheduler1coverbuddycountry2 = []
                            for x in DSscheduler1coverbuddy2:
                                if x != None:
                                    DSscheduler1coverbuddycountry2.append(x[0:4])
                            
                            #getting the country for each DS, up to 11 DS to cover
                            #we will create a list with all countries in the same order as the DS are displayed in the list
                            countrylistbuddy2 = []
                            for DSforbuddy in DSscheduler1coverbuddycountry2:
                                wconfigStations= wconfig.worksheets[1]
                                rangeconfig = wconfigStations._cells_by_col(1,1,2,10000)
                                try:    
                                    for row in rangeconfig:
                                        for cell in row:
                                            if (cell.value == DSforbuddy):
                                                countrylistbuddy2.append(wconfigStations.cell(row=cell.row, column=2).value)
                                except Exception: 
                                    pass
                            
                            
                            #colour formatting and displaying it
                            global count004
                            count004=5
                            for x,y in zip(countrylistbuddy2,DSscheduler1coverbuddy2):
                                print(x)
                                #print(y)
                                if x == "UK":
                                    colour="orange"
                                if x == "FR":
                                    colour="cyan"
                                if x == "DE":
                                    colour="yellow"
                                if x == "ES":
                                    colour="magenta"
                                if x == "IT":
                                    colour="green"
                                if x == "BE":
                                    colour = "SkyBlue4"
                                if x == "NL":
                                    colour = "pale green"
                                if x == "AT":
                                    colour = "light yellow"
                                
                                
                                DSnodebuddy2label1 = Label(framebuddy2, text = y,padx=20, pady=5, relief=GROOVE, width=8, background= colour)
                                DSnodebuddy2label1.grid(row=(count004), column=3, columnspan=1, pady=10, padx=10)
                                count004 +=1
                            


                            ##############   Displaying stations for the 2nd scheduler covering 2nd buddy ################
                            #extractng only the first 4 digits of DS node (removing SD)
                            DSscheduler2coverbuddycountry2 = []
                            for x in DSscheduler2coverbuddy2:
                                if x != None:
                                    DSscheduler2coverbuddycountry2.append(x[0:4])
                            
                            #getting the country for each DS, up to 11 DS to cover
                            #we will create a list with all countries in the same order as the DS are displayed in the list
                            countrylist2buddy2 = []
                            for DSforbuddy in DSscheduler2coverbuddycountry2:
                                wconfigStations= wconfig.worksheets[1]
                                rangeconfig = wconfigStations._cells_by_col(1,1,2,10000)
                                try:    
                                    for row in rangeconfig:
                                        for cell in row:
                                            if (cell.value == DSforbuddy):
                                                countrylist2buddy2.append(wconfigStations.cell(row=cell.row, column=2).value)
                                except Exception: 
                                    pass
                            
                            
                            #colour formatting and displaying it
                            global count005
                            count005 = 5
                            for x,y in zip(countrylist2buddy2,DSscheduler2coverbuddy2):
                                #print(x)
                                #print(y)
                                if x == "UK":
                                    colour="orange"
                                if x == "FR":
                                    colour="cyan"
                                if x == "DE":
                                    colour="yellow"
                                if x == "ES":
                                    colour="magenta"
                                if x == "IT":
                                    colour="green"
                                if x == "BE":
                                    colour = "SkyBlue4"
                                if x == "NL":
                                    colour = "pale green"
                                if x == "AT":
                                    colour = "light yellow"
                                
                                
                                DSnodebuddy2label2 = Label(framebuddy2, text = y,padx=20, pady=5, relief=GROOVE, width=8, background= colour)
                                DSnodebuddy2label2.grid(row=(count005), column=4, columnspan=1, pady=10, padx=10)
                                count005 +=1
                            ##############   Displaying stations for the 3rd scheduler covering 2nd buddy ################
                            #extractng only the first 4 digits of DS node (removing SD)
                            DSscheduler3coverbuddycountry2 = []
                            for x in DSscheduler3coverbuddy2:
                                if x != None:
                                    DSscheduler3coverbuddycountry2.append(x[0:4])
                            
                            #getting the country for each DS, up to 11 DS to cover
                            #we will create a list with all countries in the same order as the DS are displayed in the list
                            countrylist3buddy2 = []
                            for DSforbuddy in DSscheduler3coverbuddycountry2:
                                wconfigStations= wconfig.worksheets[1]
                                rangeconfig = wconfigStations._cells_by_col(1,1,2,10000)
                                try:    
                                    for row in rangeconfig:
                                        for cell in row:
                                            if (cell.value == DSforbuddy):
                                                countrylist3buddy2.append(wconfigStations.cell(row=cell.row, column=2).value)
                                except Exception: 
                                    pass
                            
                            #colour formatting and displaying it
                            global count006
                            count006 = 5
                            for x,y in zip(countrylist3buddy2,DSscheduler3coverbuddy2):
                                #print(x)
                                #print(y)
                                if x == "UK":
                                    colour="orange"
                                if x == "FR":
                                    colour="cyan"
                                if x == "DE":
                                    colour="yellow"
                                if x == "ES":
                                    colour="magenta"
                                if x == "IT":
                                    colour="green"
                                if x == "BE":
                                    colour = "SkyBlue4"
                                if x == "NL":
                                    colour = "pale green"
                                if x == "AT":
                                    colour = "light yellow"
                                
                                
                                DSnodebuddy2label2 = Label(framebuddy2, text = y,padx=20, pady=5, relief=GROOVE, width=8, background= colour)
                                DSnodebuddy2label2.grid(row=(count006), column=5, columnspan=1, pady=10, padx=10)
                                count006 +=1
                                                    
        ######################################################################################
            #searching for the DS of the third buddy 
        ######################################################################################################################################
                
            #getting the DS for 2nd buddy and appending to a list
            DSscheduler1coverbuddy3 = []
            DSscheduler2coverbuddy3 = []
            DSscheduler3coverbuddy3 = []

            ws = wb.worksheets[0] #sheet number 1
            range = ws.iter_rows()
            for row in range:
                    for cell in row:
                        if (cell.value == search_strbuddy3):
                            DSnode1buddy3 = ws.cell(row=cell.row, column=4).value #get the value from column 4 and the row where DS has been found
                            DSnode2buddy3 = ws.cell(row=cell.row, column=5).value
                            DSnode3buddy3 = ws.cell(row=cell.row, column=6).value
                            DSnode4buddy3 = ws.cell(row=cell.row, column=7).value
                            DSnode5buddy3 = ws.cell(row=cell.row, column=8).value
                            DSnode6buddy3 = ws.cell(row=cell.row, column=9).value
                            DSnode7buddy3 = ws.cell(row=cell.row, column=10).value
                            DSnode8buddy3 = ws.cell(row=cell.row, column=11).value
                            DSnode9buddy3 = ws.cell(row=cell.row, column=12).value
                            DSnode10buddy3 = ws.cell(row=cell.row, column=13).value
                            DSnode11buddy3 = ws.cell(row=cell.row, column=14).value
                            DSnode12buddy3 = ws.cell(row=cell.row, column=15).value
                            DSnode13buddy3 = ws.cell(row=cell.row, column=16).value
                            DSnode14buddy3 = ws.cell(row=cell.row, column=17).value
                            DSnode15buddy3 = ws.cell(row=cell.row, column=18).value
                            DSnode16buddy3 = ws.cell(row=cell.row, column=19).value
                            DSnode17buddy3 = ws.cell(row=cell.row, column=20).value
                            DSnode18buddy3 = ws.cell(row=cell.row, column=21).value
                            DSnode19buddy3 = ws.cell(row=cell.row, column=22).value
                            DSnode20buddy3 = ws.cell(row=cell.row, column=23).value
                            DSnode21buddy3 = ws.cell(row=cell.row, column=24).value
                            DSnode22buddy3 = ws.cell(row=cell.row, column=25).value #up to 22 DS to be split

                            #append values in list and separate which DS a buddy will cover and which ones will go for the other buddy
                            DSscheduler1coverbuddy3.extend((DSnode1buddy3,DSnode4buddy3,DSnode7buddy3,DSnode10buddy3,DSnode13buddy3,DSnode16buddy3,DSnode19buddy3,DSnode22buddy3))
                            DSscheduler2coverbuddy3.extend((DSnode2buddy3,DSnode5buddy3,DSnode8buddy3,DSnode11buddy3,DSnode14buddy3,DSnode17buddy3,DSnode20buddy3))
                            DSscheduler3coverbuddy3.extend((DSnode3buddy3,DSnode6buddy3,DSnode9buddy3,DSnode12buddy3,DSnode15buddy3,DSnode18buddy3,DSnode21buddy3))

                            #getting the country for each DS. Depending on the country it will be displayed with one colour or another       
                            shiftconfiglocation = r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\TaskList\TasklistGeneration\\' + dateStr2 + '\\shift_config.xlsx'
                            wconfig = openpyxl.load_workbook(shiftconfiglocation)
                            wconfigStations= wconfig.worksheets[1]
                            rangeconfig = wconfigStations._cells_by_col(1,1,2,10000)
                            

                            ##############   Displaying stations for the 1st scheduler covering 3rd buddy ################
                            #extractng only the first 4 digits of DS node (removing SD in some DS)
                            DSscheduler1coverbuddycountry3 = []
                            for x in DSscheduler1coverbuddy3:
                                if x != None:
                                    DSscheduler1coverbuddycountry3.append(x[0:4])
                            
                            #getting the country for each DS, up to 11 DS to cover
                            #we will create a list with all countries in the same order as the DS are displayed in the list
                            countrylistbuddy3 = []
                            for DSforbuddy in DSscheduler1coverbuddycountry3:
                                wconfigStations= wconfig.worksheets[1]
                                rangeconfig = wconfigStations._cells_by_col(1,1,2,10000)
                                try:    
                                    for row in rangeconfig:
                                        for cell in row:
                                            if (cell.value == DSforbuddy):
                                                countrylistbuddy3.append(wconfigStations.cell(row=cell.row, column=2).value)
                                except Exception: 
                                    pass
                            
                            
                            #colour formatting and displaying it
                            global count007
                            count007=5
                            for x,y in zip(countrylistbuddy3,DSscheduler1coverbuddy3):
                                print(x)
                                #print(y)
                                if x == "UK":
                                    colour="orange"
                                if x == "FR":
                                    colour="cyan"
                                if x == "DE":
                                    colour="yellow"
                                if x == "ES":
                                    colour="magenta"
                                if x == "IT":
                                    colour="green"
                                if x == "BE":
                                    colour = "SkyBlue4"
                                if x == "NL":
                                    colour = "pale green"
                                if x == "AT":
                                    colour = "light yellow"
                                
                                
                                DSnodebuddy3label1 = Label(framebuddy3, text = y,padx=20, pady=5, relief=GROOVE, width=8, background= colour)
                                DSnodebuddy3label1.grid(row=(count007), column=6, columnspan=1, pady=10, padx=10)
                                count007 +=1
                            


                            ##############   Displaying stations for the 2nd scheduler covering 3rd buddy ################
                            #extractng only the first 4 digits of DS node (removing SD)
                            DSscheduler2coverbuddycountry3 = []
                            for x in DSscheduler2coverbuddy3:
                                if x != None:
                                    DSscheduler2coverbuddycountry3.append(x[0:4])
                            
                            #getting the country for each DS, up to 11 DS to cover
                            #we will create a list with all countries in the same order as the DS are displayed in the list
                            countrylist2buddy3 = []
                            for DSforbuddy in DSscheduler2coverbuddycountry3:
                                wconfigStations= wconfig.worksheets[1]
                                rangeconfig = wconfigStations._cells_by_col(1,1,2,10000)
                                try:    
                                    for row in rangeconfig:
                                        for cell in row:
                                            if (cell.value == DSforbuddy):
                                                countrylist2buddy3.append(wconfigStations.cell(row=cell.row, column=2).value)
                                except Exception: 
                                    pass
                            
                            
                            
                            #colour formatting and displaying it
                            global count008
                            count008 = 5
                            for x,y in zip(countrylist2buddy3,DSscheduler2coverbuddy3):
                                #print(x)
                                #print(y)
                                if x == "UK":
                                    colour="orange"
                                if x == "FR":
                                    colour="cyan"
                                if x == "DE":
                                    colour="yellow"
                                if x == "ES":
                                    colour="magenta"
                                if x == "IT":
                                    colour="green"
                                if x == "BE":
                                    colour = "SkyBlue4"
                                if x == "NL":
                                    colour = "pale green"
                                if x == "AT":
                                    colour = "light yellow"
                                
                                
                                DSnodebuddy3label2 = Label(framebuddy3, text = y,padx=20, pady=5, relief=GROOVE, width=8, background= colour)
                                DSnodebuddy3label2.grid(row=(count008), column=7, columnspan=1, pady=10, padx=10)
                                count008 +=1
                            
                            ##############   Displaying stations for the 3rd scheduler covering 3rd buddy ################
                            #extractng only the first 4 digits of DS node (removing SD)
                            DSscheduler3coverbuddycountry3 = []
                            for x in DSscheduler3coverbuddy3:
                                if x != None:
                                    DSscheduler3coverbuddycountry3.append(x[0:4])
                            
                            #getting the country for each DS, up to 11 DS to cover
                            #we will create a list with all countries in the same order as the DS are displayed in the list
                            countrylist3buddy3 = []
                            for DSforbuddy in DSscheduler3coverbuddycountry3:
                                wconfigStations= wconfig.worksheets[1]
                                rangeconfig = wconfigStations._cells_by_col(1,1,2,10000)
                                try:    
                                    for row in rangeconfig:
                                        for cell in row:
                                            if (cell.value == DSforbuddy):
                                                countrylist3buddy3.append(wconfigStations.cell(row=cell.row, column=2).value)
                                except Exception: 
                                    pass
                            
                            
                            
                            #colour formatting and displaying it
                            global count009
                            count009 = 5
                            for x,y in zip(countrylist3buddy3,DSscheduler3coverbuddy3):
                                #print(x)
                                #print(y)
                                if x == "UK":
                                    colour="orange"
                                if x == "FR":
                                    colour="cyan"
                                if x == "DE":
                                    colour="yellow"
                                if x == "ES":
                                    colour="magenta"
                                if x == "IT":
                                    colour="green"
                                if x == "BE":
                                    colour = "SkyBlue4"
                                if x == "NL":
                                    colour = "pale green"
                                if x == "AT":
                                    colour = "light yellow"
                                
                                
                                DSnodebuddy3label2 = Label(framebuddy3, text = y,padx=20, pady=5, relief=GROOVE, width=8, background= colour)
                                DSnodebuddy3label2.grid(row=(count009), column=8, columnspan=1, pady=10, padx=10)
                                count009 +=1
            
            
            
                                                        
            
            ######################################################################################
            #searching for the DS of the fourth buddy 
        ######################################################################################################################################
                
            #getting the DS for 2nd buddy and appending to a list
            DSscheduler1coverbuddy4 = []
            DSscheduler2coverbuddy4 = []
            DSscheduler3coverbuddy4 = []

            ws = wb.worksheets[0] #sheet number 1
            range = ws.iter_rows()
            for row in range:
                    for cell in row:
                        if (cell.value == search_strbuddy4):
                            DSnode1buddy4 = ws.cell(row=cell.row, column=4).value #get the value from column 4 and the row where DS has been found
                            DSnode2buddy4 = ws.cell(row=cell.row, column=5).value
                            DSnode3buddy4 = ws.cell(row=cell.row, column=6).value
                            DSnode4buddy4 = ws.cell(row=cell.row, column=7).value
                            DSnode5buddy4 = ws.cell(row=cell.row, column=8).value
                            DSnode6buddy4 = ws.cell(row=cell.row, column=9).value
                            DSnode7buddy4 = ws.cell(row=cell.row, column=10).value
                            DSnode8buddy4 = ws.cell(row=cell.row, column=11).value
                            DSnode9buddy4 = ws.cell(row=cell.row, column=12).value
                            DSnode10buddy4 = ws.cell(row=cell.row, column=13).value
                            DSnode11buddy4 = ws.cell(row=cell.row, column=14).value
                            DSnode12buddy4 = ws.cell(row=cell.row, column=15).value
                            DSnode13buddy4 = ws.cell(row=cell.row, column=16).value
                            DSnode14buddy4 = ws.cell(row=cell.row, column=17).value
                            DSnode15buddy4 = ws.cell(row=cell.row, column=18).value
                            DSnode16buddy4 = ws.cell(row=cell.row, column=19).value
                            DSnode17buddy4 = ws.cell(row=cell.row, column=20).value
                            DSnode18buddy4 = ws.cell(row=cell.row, column=21).value
                            DSnode19buddy4 = ws.cell(row=cell.row, column=22).value
                            DSnode20buddy4 = ws.cell(row=cell.row, column=23).value
                            DSnode21buddy4 = ws.cell(row=cell.row, column=24).value
                            DSnode22buddy4 = ws.cell(row=cell.row, column=25).value #up to 22 DS to be split

                            #append values in list and separate which DS a buddy will cover and which ones will go for the other buddy
                            DSscheduler1coverbuddy4.extend((DSnode1buddy4,DSnode4buddy4,DSnode7buddy4,DSnode10buddy4,DSnode13buddy4,DSnode16buddy4,DSnode19buddy4,DSnode22buddy4))
                            DSscheduler2coverbuddy4.extend((DSnode2buddy4,DSnode5buddy4,DSnode8buddy4,DSnode11buddy4,DSnode14buddy4,DSnode17buddy4,DSnode20buddy4))
                            DSscheduler3coverbuddy4.extend((DSnode3buddy4,DSnode6buddy4,DSnode9buddy4,DSnode12buddy4,DSnode15buddy4,DSnode18buddy4,DSnode21buddy4))

                            #getting the country for each DS. Depending on the country it will be displayed with one colour or another       
                            shiftconfiglocation = r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\TaskList\TasklistGeneration\\' + dateStr2 + '\\shift_config.xlsx'
                            wconfig = openpyxl.load_workbook(shiftconfiglocation)
                            wconfigStations= wconfig.worksheets[1]
                            rangeconfig = wconfigStations._cells_by_col(1,1,2,10000)
                            

                            ##############   Displaying stations for the 1st scheduler covering 3rd buddy ################
                            #extractng only the first 4 digits of DS node (removing SD in some DS)
                            DSscheduler1coverbuddycountry4 = []
                            for x in DSscheduler1coverbuddy4:
                                if x != None:
                                    DSscheduler1coverbuddycountry4.append(x[0:4])
                            
                            #getting the country for each DS, up to 11 DS to cover
                            #we will create a list with all countries in the same order as the DS are displayed in the list
                            countrylistbuddy4 = []
                            for DSforbuddy in DSscheduler1coverbuddycountry4:
                                wconfigStations= wconfig.worksheets[1]
                                rangeconfig = wconfigStations._cells_by_col(1,1,2,10000)
                                try:    
                                    for row in rangeconfig:
                                        for cell in row:
                                            if (cell.value == DSforbuddy):
                                                countrylistbuddy4.append(wconfigStations.cell(row=cell.row, column=2).value)
                                except Exception: 
                                    pass
                            
                            
                            
                            #colour formatting and displaying it
                            global count010
                            count010=5
                            for x,y in zip(countrylistbuddy4,DSscheduler1coverbuddy4):
                                print(x)
                                #print(y)
                                if x == "UK":
                                    colour="orange"
                                if x == "FR":
                                    colour="cyan"
                                if x == "DE":
                                    colour="yellow"
                                if x == "ES":
                                    colour="magenta"
                                if x == "IT":
                                    colour="green"
                                if x == "BE":
                                    colour = "SkyBlue4"
                                if x == "NL":
                                    colour = "pale green"
                                if x == "AT":
                                    colour = "light yellow"
                                
                                
                                DSnodebuddy4label1 = Label(framebuddy4, text = y,padx=20, pady=5, relief=GROOVE, width=8, background= colour)
                                DSnodebuddy4label1.grid(row=(count010), column=9, columnspan=1, pady=10, padx=10)
                                count010 +=1
                            


                            ##############   Displaying stations for the 2nd scheduler covering 4th buddy ################
                            #extractng only the first 4 digits of DS node (removing SD)
                            DSscheduler2coverbuddycountry4 = []
                            for x in DSscheduler2coverbuddy4:
                                if x != None:
                                    DSscheduler2coverbuddycountry4.append(x[0:4])
                            
                            #getting the country for each DS, up to 11 DS to cover
                            #we will create a list with all countries in the same order as the DS are displayed in the list
                            countrylist2buddy4 = []
                            for DSforbuddy in DSscheduler2coverbuddycountry4:
                                wconfigStations= wconfig.worksheets[1]
                                rangeconfig = wconfigStations._cells_by_col(1,1,2,10000)
                                try:    
                                    for row in rangeconfig:
                                        for cell in row:
                                            if (cell.value == DSforbuddy):
                                                countrylist2buddy4.append(wconfigStations.cell(row=cell.row, column=2).value)
                                except Exception: 
                                    pass
                            
                            
                            
                            #colour formatting and displaying it
                            global count011
                            count011 = 5
                            for x,y in zip(countrylist2buddy4,DSscheduler2coverbuddy4):
                                #print(x)
                                #print(y)
                                if x == "UK":
                                    colour="orange"
                                if x == "FR":
                                    colour="cyan"
                                if x == "DE":
                                    colour="yellow"
                                if x == "ES":
                                    colour="magenta"
                                if x == "IT":
                                    colour="green"
                                if x == "BE":
                                    colour = "SkyBlue4"
                                if x == "NL":
                                    colour = "pale green"
                                if x == "AT":
                                    colour = "light yellow"
                                
                                
                                DSnodebuddy4label2 = Label(framebuddy4, text = y,padx=20, pady=5, relief=GROOVE, width=8, background= colour)
                                DSnodebuddy4label2.grid(row=(count011), column=10, columnspan=1, pady=10, padx=10)
                                count011 +=1
                            
                            ##############   Displaying stations for the 3rd scheduler covering 4th buddy ################
                            #extractng only the first 4 digits of DS node (removing SD)
                            DSscheduler3coverbuddycountry4 = []
                            for x in DSscheduler3coverbuddy4:
                                if x != None:
                                    DSscheduler3coverbuddycountry4.append(x[0:4])
                            
                            #getting the country for each DS, up to 11 DS to cover
                            #we will create a list with all countries in the same order as the DS are displayed in the list
                            countrylist3buddy4 = []
                            for DSforbuddy in DSscheduler3coverbuddycountry4:
                                wconfigStations= wconfig.worksheets[1]
                                rangeconfig = wconfigStations._cells_by_col(1,1,2,10000)
                                try:    
                                    for row in rangeconfig:
                                        for cell in row:
                                            if (cell.value == DSforbuddy):
                                                countrylist3buddy4.append(wconfigStations.cell(row=cell.row, column=2).value)
                                except Exception: 
                                    pass
                            
                            
                            
                            #colour formatting and displaying it
                            global count012
                            count012 = 5
                            for x,y in zip(countrylist3buddy4,DSscheduler3coverbuddy4):
                                #print(x)
                                #print(y)
                                if x == "UK":
                                    colour="orange"
                                if x == "FR":
                                    colour="cyan"
                                if x == "DE":
                                    colour="yellow"
                                if x == "ES":
                                    colour="magenta"
                                if x == "IT":
                                    colour="green"
                                if x == "BE":
                                    colour = "SkyBlue4"
                                if x == "NL":
                                    colour = "pale green"
                                if x == "AT":
                                    colour = "light yellow"
                                
                                
                                DSnodebuddy4label2 = Label(framebuddy4, text = y,padx=20, pady=5, relief=GROOVE, width=8, background= colour)
                                DSnodebuddy4label2.grid(row=(count012), column=11, columnspan=1, pady=10, padx=10)
                                count012 +=1
            
            
            
            # #adding button to import DS        
            # def ImportDSbuddy4scheduler1():
            #     answer = tkinter.messagebox.askokcancel("IMPORTANT", "Make sure you clear all your notifications in Amazon Chime before proceeding")
            #     if answer:
            #         for x in DSscheduler1coverbuddycountry4:       
            #             if x != None:
            #                 if autoit.win_exists("Amazon Chime"):
            #                     autoit.win_activate("Amazon Chime")
            #                     autoit.win_wait("Amazon Chime")
            #                     time.sleep(0.8)
            #                     autoit.send('^f')
            #                     time.sleep(0.8)
            #                     autoit.send('^a')
            #                     time.sleep(0.8)
            #                     autoit.send('{BACKSPACE}')
            #                     time.sleep(0.8)
            #                     autoit.send(x)
            #                     time.sleep(0.8)
            #                     autoit.send('{TAB}')
            #                     time.sleep(0.8)
            #                     autoit.send('{TAB}')
            #                     time.sleep(0.8)
            #                     autoit.send('{ENTER}')
            #                     time.sleep(0.8)
            #                 else:
            #                     print("error")
            #         tkinter.messagebox.showinfo('Stations Added to Chime', 'All station chat rooms for the selected user have now been added to chime!')
            
            # def ImportDSbuddy4scheduler2():
            #     answer = tkinter.messagebox.askokcancel("IMPORTANT", "Make sure you clear all your notifications in Amazon Chime before proceeding")
            #     if answer:
            #         for x in DSscheduler2coverbuddycountry4:       
            #             if x != None:
            #                 if autoit.win_exists("Amazon Chime"):
            #                     autoit.win_activate("Amazon Chime")
            #                     autoit.win_wait("Amazon Chime")
            #                     time.sleep(0.8)
            #                     autoit.send('^f')
            #                     time.sleep(0.8)
            #                     autoit.send('^a')
            #                     time.sleep(0.8)
            #                     autoit.send('{BACKSPACE}')
            #                     time.sleep(0.8)
            #                     autoit.send(x)
            #                     time.sleep(0.8)
            #                     autoit.send('{TAB}')
            #                     time.sleep(0.8)
            #                     autoit.send('{TAB}')
            #                     time.sleep(0.8)
            #                     autoit.send('{ENTER}')
            #                     time.sleep(0.8)
            #                 else:
            #                     print("error")
            #         tkinter.messagebox.showinfo('Stations Added to Chime', 'All station chat rooms for the selected user have now been added to chime!')
            
            # def ImportDSbuddy4scheduler3():
            #     answer = tkinter.messagebox.askokcancel("IMPORTANT", "Make sure you clear all your notifications in Amazon Chime before proceeding")
            #     if answer:
            #         for x in DSscheduler3coverbuddycountry4:       
            #             if x != None:
            #                 if autoit.win_exists("Amazon Chime"):
            #                     autoit.win_activate("Amazon Chime")
            #                     autoit.win_wait("Amazon Chime")
            #                     time.sleep(0.8)
            #                     autoit.send('^f')
            #                     time.sleep(0.8)
            #                     autoit.send('^a')
            #                     time.sleep(0.8)
            #                     autoit.send('{BACKSPACE}')
            #                     time.sleep(0.8)
            #                     autoit.send(x)
            #                     time.sleep(0.8)
            #                     autoit.send('{TAB}')
            #                     time.sleep(0.8)
            #                     autoit.send('{TAB}')
            #                     time.sleep(0.8)
            #                     autoit.send('{ENTER}')
            #                     time.sleep(0.8)
            #                 else:
            #                     print("error")
            #         tkinter.messagebox.showinfo('Stations Added to Chime', 'All station chat rooms for the selected user have now been added to chime!')
            
            # #adding button to import DS
            # ButtonImportDSCover4Scheduler1 = Button(framebuddy4, text = "Import Stations", padx=20, pady=5, relief=GROOVE, width=7, command=ImportDSbuddy4scheduler1, bg = "white", font=("Helvetica", 10))
            # ButtonImportDSCover4Scheduler1.grid(row=count010, column=9, columnspan=1, pady=10, padx=10)
            # ButtonImportDSCover4Scheduler2 = Button(framebuddy4, text = "Import Stations", padx=20, pady=5, relief=GROOVE, width=7, command=ImportDSbuddy4scheduler2, bg = "white", font=("Helvetica", 10))
            # ButtonImportDSCover4Scheduler2.grid(row=count011, column=10, columnspan=1, pady=10, padx=10)
            # ButtonImportDSCover4Scheduler3 = Button(framebuddy4, text = "Import Stations", padx=20, pady=5, relief=GROOVE, width=7, command=ImportDSbuddy4scheduler3, bg = "white", font=("Helvetica", 10))
            # ButtonImportDSCover4Scheduler3.grid(row=count012, column=11, columnspan=1, pady=10, padx=10)    

            #getting the breaks and adding to table
            
            os.chdir(r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\TaskList\TasklistGeneration\\' + dateStr2 )
            try:
                breakcsv = csv.reader(open("breakinfo.csv","r"))
                for row in breakcsv:
                    if search_strbuddy1 == row[0]:
                        Breaktime = row
                        BreakScheduler1 = Breaktime[1]
            except:
                pass

            try:
                breakcsv = csv.reader(open("breakinfo.csv","r"))
                for row in breakcsv:
                    if search_strbuddy2 == row[0]:
                        Breaktime2 = row
                        BreakScheduler2 = Breaktime2[1]
            except:
                pass
            try:
                breakcsv = csv.reader(open("breakinfo.csv","r"))
                for row in breakcsv:
                    if search_strbuddy3 == row[0]:
                        Breaktime3 = row
                        BreakScheduler3 = Breaktime3[1]
            except:
                pass
            try:
                breakcsv = csv.reader(open("breakinfo.csv","r"))
                for row in breakcsv:
                    if search_strbuddy4 == row[0]:
                        Breaktime4 = row
                        BreakScheduler4 = Breaktime4[1]
            except:
                pass
            
        #adding labels
            try:
                BreakScheduler1label = Label(framebuddy1,  text= BreakScheduler1, justify=CENTER, padx=20, pady=1, fg="red", font=("Helvetica", 10, 'bold'))
                BreakScheduler1label.grid(row=3, column=0, columnspan=3, pady=5, padx=10)
            except:
                pass
            try:
                BreakScheduler2label = Label(framebuddy2,  text= BreakScheduler2, justify=CENTER, padx=20, pady=1,fg="red", font=("Helvetica", 10, 'bold'))
                BreakScheduler2label.grid(row=3, column=3, columnspan=3, pady=5, padx=10)
            except:
                pass
            try:
                BreakScheduler3label = Label(framebuddy3,  text= BreakScheduler3, justify=CENTER, padx=20, pady=1,fg="red", font=("Helvetica", 10, 'bold'))
                BreakScheduler3label.grid(row=3, column=6, columnspan=3, pady=5, padx=10)
            except:
                pass
            try:
                BreakScheduler4label = Label(framebuddy4,  text= BreakScheduler4, justify=CENTER, padx=20, pady=1,fg="red", font=("Helvetica", 10, 'bold'))
                BreakScheduler4label.grid(row=3, column=9, columnspan=3, pady=5, padx=10)
            except:
                pass

            #get the confirmation and adding to table
            
            os.chdir(r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\TaskList\TasklistGeneration\\' + dateStr2 )
            try:
                breakcsv = csv.reader(open("breakinfoconfirmation.csv","r"))
                for row in breakcsv:
                    if search_strbuddy1 == row[0]:
                        confirmation = row
                        ConfirmationScheduler1 = confirmation[1]
            except:
                pass

            try:
                breakcsv = csv.reader(open("breakinfoconfirmation.csv","r"))
                for row in breakcsv:
                    if search_strbuddy2 == row[0]:
                        confirmation2 = row
                        ConfirmationScheduler2 = confirmation2[1]
            except:
                pass
            try:
                breakcsv = csv.reader(open("breakinfoconfirmation.csv","r"))
                for row in breakcsv:
                    if search_strbuddy3 == row[0]:
                        confirmation3 = row
                        ConfirmationScheduler3 = confirmation3[1]
            except:
                pass
            try:
                breakcsv = csv.reader(open("breakinfoconfirmation.csv","r"))
                for row in breakcsv:
                    if search_strbuddy4 == row[0]:
                        confirmation4 = row
                        ConfirmationScheduler4 = confirmation4[1]
            except:
                pass

            #adding labels
            try:
                if ConfirmationScheduler1 != None:
                    ConfirmationScheduler1label = Label(framebuddy1, text= search_strbuddy1 + " confirms covering breaks", foreground= "green", relief=GROOVE, justify=CENTER, padx=20, pady=5)
                    ConfirmationScheduler1label.grid(column = 0, row= count001+2, pady=10, padx=10, columnspan=3)
                    print("confirmed")
            except:
                pass
            try:
                if ConfirmationScheduler2 != None:
                    ConfirmationScheduler2label = Label(framebuddy2, text= search_strbuddy2 + " confirms covering breaks", foreground= "green", relief=GROOVE, justify=CENTER, padx=20, pady=5)
                    ConfirmationScheduler2label.grid(column = 3, row= count004+2, pady=10, padx=10, columnspan=3)
                    print("confirmed")
            except:
                pass
            try:
                if ConfirmationScheduler3 != None:
                    ConfirmationScheduler3label = Label(framebuddy3, text= search_strbuddy3 + " confirms covering breaks", foreground= "green", relief=GROOVE, justify=CENTER, padx=20, pady=5)
                    ConfirmationScheduler3label.grid(column = 6, row= count007+2, pady=10, padx=10, columnspan=3)
                    print("confirmed")
            except:
                pass
            try:
                if ConfirmationScheduler4 != None:
                    ConfirmationScheduler4label = Label(framebuddy4, text= search_strbuddy4 + " confirms covering breaks", foreground= "green", relief=GROOVE, justify=CENTER, padx=20, pady=5)
                    ConfirmationScheduler4label.grid(column = 9, row= count010+2, pady=10, padx=10, columnspan=3)
                    print("confirmed")
            except:
                pass

            #getting the extracovers and adding to table
            breakcsv = csv.reader(open("extracover.csv","r"))
            for row in breakcsv:
                if search_strbuddy1 == row[0]:
                    extracoverDS = row
                    extra1cover = extracoverDS[1] 
                    extra2cover = extracoverDS[2] 
                    print(extra1cover)
                    
                    extra1coverlabelinfo = Label(framebuddy1, text= search_strbuddy1 + " has extra cover", relief=GROOVE, justify=CENTER, padx=20, pady=5, width=50)
                    extra1coverlabelinfo.grid(column = 0, row= count001+3, pady=10, padx=10, columnspan=3)
                    if extra1cover != "":
                        extra1coverlabel = Label(framebuddy1, text= extra1cover, relief=GROOVE, justify=CENTER, padx=20, pady=5)
                        extra1coverlabel.grid(column = 1, row= count001+4, pady=10, padx=10, columnspan=1)
                        DSscheduler2coverbuddycountry1.append(extra1cover)

                    if extra2cover != "":
                        extra2coverlabel = Label(framebuddy1, text= extra2cover, relief=GROOVE, justify=CENTER, padx=20, pady=5)
                        extra2coverlabel.grid(column = 2, row= count001+4, pady=10, padx=10, columnspan=1)
                        DSscheduler3coverbuddycountry1.append(extra2cover)
            
            breakcsv = csv.reader(open("extracover.csv","r"))
            for row in breakcsv:
                if search_strbuddy2 == row[0]:
                    extracoverDS2 = row
                    extra1cover2 = extracoverDS2[1] 
                    extra2cover2 = extracoverDS2[2] 
                    
                    extra1coverlabelinfo = Label(framebuddy2, text= search_strbuddy2 + " has extra cover", relief=GROOVE, justify=CENTER, padx=20, pady=5, width=50)
                    extra1coverlabelinfo.grid(column = 3, row= count004+3, pady=10, padx=10, columnspan=3)
                    if extra1cover2 != "":
                        extra1coverlabel2 = Label(framebuddy2, text= extra1cover2, relief=GROOVE, justify=CENTER, padx=20, pady=5)
                        extra1coverlabel2.grid(column = 4, row= count004+4, pady=10, padx=10, columnspan=1) 
                        DSscheduler2coverbuddycountry2.append(extra1cover2)

                    if extra2cover2 != "":
                        extra2coverlabel2 = Label(framebuddy2, text= extra2cover2, relief=GROOVE, justify=CENTER, padx=20, pady=5)
                        extra2coverlabel2.grid(column = 5, row= count004+4, pady=10, padx=10, columnspan=1)
                        DSscheduler3coverbuddycountry2.append(extra2cover2)
                
                    
            
            breakcsv = csv.reader(open("extracover.csv","r"))
            for row in breakcsv:
                if search_strbuddy3 == row[0]:
                    extracoverDS3 = row
                    extra1cover3 = extracoverDS3[1] 
                    extra2cover3 = extracoverDS3[2]
                    extra1coverlabelinfo = Label(framebuddy3, text= search_strbuddy3 + " has extra cover", relief=GROOVE, justify=CENTER, padx=20, pady=5, width=50)
                    extra1coverlabelinfo.grid(column = 6, row= count007+3, pady=10, padx=10, columnspan=3)
                    if extra1cover3 != "":
                        extra1coverlabel3 = Label(framebuddy3, text= extra1cover3, relief=GROOVE, justify=CENTER, padx=20, pady=5)
                        extra1coverlabel3.grid(column = 7, row= count007+4, pady=10, padx=10, columnspan=1) 
                        DSscheduler2coverbuddycountry3.append(extra1cover3)
                    if extra2cover3 != "":
                        extra2coverlabel3 = Label(framebuddy3, text= extra2cover3, relief=GROOVE, justify=CENTER, padx=20, pady=5)
                        extra2coverlabel3.grid(column = 8, row= count007+4, pady=10, padx=10, columnspan=1)
                        DSscheduler3coverbuddycountry3.append(extra2cover3)
            
            breakcsv = csv.reader(open("extracover.csv","r"))
            for row in breakcsv:
                if search_strbuddy4 == row[0]:
                    extracoverDS4 = row
                    extra1cover4 = extracoverDS4[1] 
                    extra2cover4 = extracoverDS4[2]
                    extra1coverlabelinfo = Label(framebuddy4, text= search_strbuddy4 + " has extra cover", relief=GROOVE, justify=CENTER, padx=20, pady=5, width=50)
                    extra1coverlabelinfo.grid(column = 9, row= count007+3, pady=10, padx=10, columnspan=3)
                    if extra1cover4 != "":
                        extra1coverlabel4 = Label(framebuddy4, text= extra1cover4, relief=GROOVE, justify=CENTER, padx=20, pady=5)
                        extra1coverlabel4.grid(column = 10, row= count007+4, pady=10, padx=10, columnspan=1) 
                        DSscheduler2coverbuddycountry4.append(extra1cover4)
                    if extra2cover4 != "":
                        extra2coverlabel4 = Label(framebuddy4, text= extra2cover4, relief=GROOVE, justify=CENTER, padx=20, pady=5)
                        extra2coverlabel4.grid(column = 11, row= count007+4, pady=10, padx=10, columnspan=1)
                        DSscheduler3coverbuddycountry4.append(extra1cover4)
        


        #activating the button when buddies confirmed
        try:
            if Buddy3 == None and Buddy4==None:
                if BreakScheduler1 and BreakScheduler2:
                    Confirm.config(state= NORMAL) 
        except:
            pass
        try:
            if Buddy3 != None and Buddy4==None:
                if BreakScheduler1 and BreakScheduler2 and BreakScheduler3 != None:
                    Confirm.config(state= NORMAL) 
        except:
            pass
        try:
            if Buddy4 != None:
                if BreakScheduler1 and BreakScheduler2 and BreakScheduler3 != None and BreakScheduler4 != None:
                    Confirm.config(state= NORMAL) 
        except:
            pass
        

        def reminder():
            os.chdir(r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\Chime Broadcaster')#webhook addresses
            answer = tkinter.messagebox.askokcancel("IMPORTANT", "By clicking OK you will be sending Webhooks to all your DS notifying you will be on break")
            if answer:
                if Buddy3 == None and Buddy4==None:
                    Sched1 = Buddy1list[1] #take the 2nd element from the buddylist
                    Sched2 = Buddy2list[1] #do the same
                    ws = wb.worksheets[0] #sheet number 1
                    range = ws.iter_rows()
                    for row in range:
                            for cell in row:
                                if (cell.value == Sched1):
                                    loginbuddy1 = ws.cell(row=cell.row, column=1).value
                    range = ws.iter_rows()
                    for row in range:
                            for cell in row:
                                if (cell.value == Sched2):
                                    loginbuddy2 = ws.cell(row=cell.row, column=1).value
                    
                        
                    if NameScheduler== Sched1:
                        for x in DSscheduler1coverbuddycountry1:       
                            if x != None:
                                webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
                                for row in webhookcsv:
                                    if (x == row[0]):
                                        webhookurlrow =  row
                                        webhookurl = webhookurlrow[1]
                                        result = False
                                        session = requests.session()
                                        data = {"Content": "/md \n\n### Scheduler on break \nHello team, **" + NameScheduler + "** is going for a break. Please note that @" + loginbuddy2 + " will be your routing POC in the meantime." }
                                        print(data)
                                        params = {'format': 'application/json'}
                                        response = session.post(webhookurl, params=params, json=data)
                                        if response.status_code == 200:
                                            result = True

                            else:
                                print("error")
                    

                    if NameScheduler== Sched2:
                        for x in DSscheduler1coverbuddycountry2:       
                            if x != None:
                                webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
                                for row in webhookcsv:
                                    if (x == row[0]):
                                        webhookurlrow =  row
                                        webhookurl = webhookurlrow[1]
                                        result = False
                                        session = requests.session()
                                        data = {"Content": "/md \n\n### Scheduler on break \nHello team, **" + NameScheduler + "** is going for a break. Please note that @" + loginbuddy1 + " will be your routing POC in the meantime." }
                                        print(data)
                                        params = {'format': 'application/json'}
                                        response = session.post(webhookurl, params=params, json=data)
                                        if response.status_code == 200:
                                            result = True

                            else:
                                print("error")
                        
                


                if Buddy3 != None and Buddy4 == None:
                    Sched1 = Buddy1list[2] #take the 3rd element from the buddylist to get just 1 name
                    Sched2 = Buddy2list[2]
                    Sched3 = Buddy3list[2]
                    ws = wb.worksheets[0] #sheet number 1
                    range = ws.iter_rows()
                    for row in range:
                            for cell in row:
                                if (cell.value == Sched1):
                                    loginbuddy1 = ws.cell(row=cell.row, column=1).value
                    range = ws.iter_rows()
                    for row in range:
                            for cell in row:
                                if (cell.value == Sched2):
                                    loginbuddy2 = ws.cell(row=cell.row, column=1).value
                    range = ws.iter_rows()
                    for row in range:
                            for cell in row:
                                if (cell.value == Sched3):
                                    loginbuddy3 = ws.cell(row=cell.row, column=1).value
                                    
                    
                    
                    if NameScheduler == Sched1:
                        for x in DSscheduler1coverbuddycountry1:       
                            if x != None:
                                webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
                                for row in webhookcsv:
                                    if (x == row[0]):
                                        webhookurlrow =  row
                                        webhookurl = webhookurlrow[1]
                                        result = False
                                        session = requests.session()
                                        data = {"Content": "/md \n\n### Scheduler on break \nHello team, **" + NameScheduler + "** is going for a break. Please note that @" + loginbuddy2 + " will be your routing POC in the meantime." }
                                        print(data)
                                        params = {'format': 'application/json'}
                                        response = session.post(webhookurl, params=params, json=data)
                                        if response.status_code == 200:
                                            result = True

                            else:
                                print("error")
                    
                        for x in DSscheduler2coverbuddycountry1:       
                            if x != None:
                                webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
                                for row in webhookcsv:
                                    if (x == row[0]):
                                        webhookurlrow =  row
                                        webhookurl = webhookurlrow[1]
                                        result = False
                                        session = requests.session()
                                        data = {"Content": "/md \n\n### Scheduler on break \nHello team, **" + NameScheduler + "** is going for a break. Please note that @" + loginbuddy3 + " will be your routing POC in the meantime." }
                                        print(data)
                                        params = {'format': 'application/json'}
                                        response = session.post(webhookurl, params=params, json=data)
                                        if response.status_code == 200:
                                            result = True
                            
                            else:
                                print("error")


                    if NameScheduler== Sched2:
                        for x in DSscheduler1coverbuddycountry2:       
                            if x != None:
                                webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
                                for row in webhookcsv:
                                    if (x == row[0]):
                                        webhookurlrow =  row
                                        webhookurl = webhookurlrow[1]
                                        result = False
                                        session = requests.session()
                                        data = {"Content": "/md \n\n### Scheduler on break \nHello team, **" + NameScheduler + "** is going for a break. Please note that @" + loginbuddy3 + " will be your routing POC in the meantime." }
                                        print(data)
                                        params = {'format': 'application/json'}
                                        response = session.post(webhookurl, params=params, json=data)
                                        if response.status_code == 200:
                                            result = True
                            
                            else:
                                print("error")
                        for x in DSscheduler2coverbuddycountry2:       
                            if x != None:
                                webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
                                for row in webhookcsv:
                                    if (x == row[0]):
                                        webhookurlrow =  row
                                        webhookurl = webhookurlrow[1]
                                        print(x + webhookurl)
                                        result = False
                                        session = requests.session()
                                        data = {"Content": "/md \n\n### Scheduler on break \nHello team, **" + NameScheduler + "** is going for a break. Please note that @" + loginbuddy1 + " will be your routing POC in the meantime." }
                                        print(data)
                                        params = {'format': 'application/json'}
                                        response = session.post(webhookurl, params=params, json=data)
                                        if response.status_code == 200:
                                            result = True
                            else:
                                print("error")
                        
                        
                    if NameScheduler== Sched3:
                        for x in DSscheduler1coverbuddycountry3:       
                            if x != None:
                                webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
                                for row in webhookcsv:
                                    if (x == row[0]):
                                        webhookurlrow =  row
                                        webhookurl = webhookurlrow[1]
                                        result = False
                                        session = requests.session()
                                        data = {"Content": "/md \n\n### Scheduler on break \nHello team, **" + NameScheduler + "** is going for a break. Please note that @" + loginbuddy1 + " will be your routing POC in the meantime." }
                                        print(data)
                                        params = {'format': 'application/json'}
                                        response = session.post(webhookurl, params=params, json=data)
                                        if response.status_code == 200:
                                            result = True
                            else:
                                print("error")
                        for x in DSscheduler2coverbuddycountry3:       
                            if x != None:
                                webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
                                for row in webhookcsv:
                                    if (x == row[0]):
                                        webhookurlrow =  row
                                        webhookurl = webhookurlrow[1]
                                        result = False
                                        session = requests.session()
                                        data = {"Content": "/md \n\n### Scheduler on break \nHello team, **" + NameScheduler + "** is going for a break. Please note that @" + loginbuddy2 + " will be your routing POC in the meantime." }
                                        print(data)
                                        params = {'format': 'application/json'}
                                        response = session.post(webhookurl, params=params, json=data)
                                        if response.status_code == 200:
                                            result = True

                            else:
                                print("error")
                        
                
                if Buddy4 != None:
                    Sched1 = Buddy1list[3] #take the 4th element from the buddylist to get just 1 name
                    Sched2 = Buddy2list[3]
                    Sched3 = Buddy3list[3]
                    Sched4 = Buddy4list[3]
                    ws = wb.worksheets[0] #sheet number 1
                    range = ws.iter_rows()
                    for row in range:
                            for cell in row:
                                if (cell.value == Sched1):
                                    loginbuddy1 = ws.cell(row=cell.row, column=1).value
                    range = ws.iter_rows()
                    for row in range:
                            for cell in row:
                                if (cell.value == Sched2):
                                    loginbuddy2 = ws.cell(row=cell.row, column=1).value
                    range = ws.iter_rows()
                    for row in range:
                            for cell in row:
                                if (cell.value == Sched3):
                                    loginbuddy3 = ws.cell(row=cell.row, column=1).value
                    range = ws.iter_rows()
                    for row in range:
                            for cell in row:
                                if (cell.value == Sched4):
                                    loginbuddy4 = ws.cell(row=cell.row, column=1).value
                    
                    if NameScheduler== Sched1:
                        for x in DSscheduler1coverbuddycountry1:       
                            if x != None:
                                webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
                                for row in webhookcsv:
                                    if (x == row[0]):
                                        webhookurlrow =  row
                                        webhookurl = webhookurlrow[1]
                                        result = False
                                        session = requests.session()
                                        data = {"Content": "/md \n\n### Scheduler on break \nHello team, **" + NameScheduler + "** is going for a break. Please note that @" + loginbuddy2 + " will be your routing POC in the meantime." }
                                        print(data)
                                        params = {'format': 'application/json'}
                                        response = session.post(webhookurl, params=params, json=data)
                                        if response.status_code == 200:
                                            result = True
                            else:
                                print("error")
                        for x in DSscheduler2coverbuddycountry1:       
                            if x != None:
                                webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
                                for row in webhookcsv:
                                    if (x == row[0]):
                                        webhookurlrow =  row
                                        webhookurl = webhookurlrow[1]
                                        result = False
                                        session = requests.session()
                                        data = {"Content": "/md \n\n### Scheduler on break \nHello team, **" + NameScheduler + "** is going for a break. Please note that @" + loginbuddy3 + " will be your routing POC in the meantime." }
                                        print(data)
                                        params = {'format': 'application/json'}
                                        response = session.post(webhookurl, params=params, json=data)
                                        if response.status_code == 200:
                                            result = True
                            else:
                                print("error")
                        
                        for x in DSscheduler3coverbuddycountry1:       
                            if x != None:
                                webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
                                for row in webhookcsv:
                                    if (x == row[0]):
                                        webhookurlrow =  row
                                        webhookurl = webhookurlrow[1]
                                        result = False
                                        session = requests.session()
                                        data = {"Content": "/md \n\n### Scheduler on break \nHello team, **" + NameScheduler + "** is going for a break. Please note that @" + loginbuddy4 + " will be your routing POC in the meantime." }
                                        print(data)
                                        params = {'format': 'application/json'}
                                        response = session.post(webhookurl, params=params, json=data)
                                        if response.status_code == 200:
                                            result = True
                            else:
                                print("error")
                        
                        
                        
                        
                        
                    
                    if NameScheduler== Sched2:
                        for x in DSscheduler1coverbuddycountry2:       
                            if x != None:
                                webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
                                for row in webhookcsv:
                                    if (x == row[0]):
                                        webhookurlrow =  row
                                        webhookurl = webhookurlrow[1]
                                        result = False
                                        session = requests.session()
                                        data = {"Content": "/md \n\n### Scheduler on break \nHello team, **" + NameScheduler + "** is going for a break. Please note that @" + loginbuddy3 + " will be your routing POC in the meantime." }
                                        print(data)
                                        params = {'format': 'application/json'}
                                        response = session.post(webhookurl, params=params, json=data)
                                        if response.status_code == 200:
                                            result = True
                            else:
                                print("error")
                        
                        for x in DSscheduler2coverbuddycountry2:       
                            if x != None:
                                webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
                                for row in webhookcsv:
                                    if (x == row[0]):
                                        webhookurlrow =  row
                                        webhookurl = webhookurlrow[1]
                                        result = False
                                        session = requests.session()
                                        data = {"Content": "/md \n\n### Scheduler on break \nHello team, **" + NameScheduler + "** is going for a break. Please note that @" + loginbuddy4 + " will be your routing POC in the meantime." }
                                        print(data)
                                        params = {'format': 'application/json'}
                                        response = session.post(webhookurl, params=params, json=data)
                                        if response.status_code == 200:
                                            result = True
                            else:
                                print("error")
                        
                        for x in DSscheduler3coverbuddycountry2:       
                            if x != None:
                                webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
                                for row in webhookcsv:
                                    if (x == row[0]):
                                        webhookurlrow =  row
                                        webhookurl = webhookurlrow[1]
                                        result = False
                                        session = requests.session()
                                        data = {"Content": "/md \n\n### Scheduler on break \nHello team, **" + NameScheduler + "** is going for a break. Please note that @" + loginbuddy1 + " will be your routing POC in the meantime." }
                                        print(data)
                                        params = {'format': 'application/json'}
                                        response = session.post(webhookurl, params=params, json=data)
                                        if response.status_code == 200:
                                            result = True
                            else:
                                print("error")
                        
                    
                        

                    if NameScheduler== Sched3:
                        for x in DSscheduler1coverbuddycountry3:       
                            if x != None:
                                webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
                                for row in webhookcsv:
                                    if (x == row[0]):
                                        webhookurlrow =  row
                                        webhookurl = webhookurlrow[1]
                                        result = False
                                        session = requests.session()
                                        data = {"Content": "/md \n\n### Scheduler on break \nHello team, **" + NameScheduler + "** is going for a break. Please note that @" + loginbuddy4 + " will be your routing POC in the meantime." }
                                        print(data)
                                        params = {'format': 'application/json'}
                                        response = session.post(webhookurl, params=params, json=data)
                                        if response.status_code == 200:
                                            result = True
                            else:
                                print("error")
                        
                        for x in DSscheduler2coverbuddycountry3:       
                            if x != None:
                                webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
                                for row in webhookcsv:
                                    if (x == row[0]):
                                        webhookurlrow =  row
                                        webhookurl = webhookurlrow[1]
                                        result = False
                                        session = requests.session()
                                        data = {"Content": "/md \n\n### Scheduler on break \nHello team, **" + NameScheduler + "** is going for a break. Please note that @" + loginbuddy1 + " will be your routing POC in the meantime." }
                                        print(data)
                                        params = {'format': 'application/json'}
                                        response = session.post(webhookurl, params=params, json=data)
                                        if response.status_code == 200:
                                            result = True
                            else:
                                print("error")
                        
                        for x in DSscheduler3coverbuddycountry3:       
                            if x != None:
                                webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
                                for row in webhookcsv:
                                    if (x == row[0]):
                                        webhookurlrow =  row
                                        webhookurl = webhookurlrow[1]
                                        result = False
                                        session = requests.session()
                                        data = {"Content": "/md \n\n### Scheduler on break \nHello team, **" + NameScheduler + "** is going for a break. Please note that @" + loginbuddy2 + " will be your routing POC in the meantime." }
                                        print(data)
                                        params = {'format': 'application/json'}
                                        response = session.post(webhookurl, params=params, json=data)
                                        if response.status_code == 200:
                                            result = True
                            else:
                                print("error")

                        
                        
                        

                    if NameScheduler== Sched4:
                        for x in DSscheduler1coverbuddycountry4:       
                            if x != None:
                                webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
                                for row in webhookcsv:
                                    if (x == row[0]):
                                        webhookurlrow =  row
                                        webhookurl = webhookurlrow[1]
                                        result = False
                                        session = requests.session()
                                        data = {"Content": "/md \n\n### Scheduler on break \nHello team, **" + NameScheduler + "** is going for a break. Please note that @" + loginbuddy1 + " will be your routing POC in the meantime." }
                                        print(data)
                                        params = {'format': 'application/json'}
                                        response = session.post(webhookurl, params=params, json=data)
                                        if response.status_code == 200:
                                            result = True
                            else:
                                print("error")
                        
                        for x in DSscheduler2coverbuddycountry4:       
                            if x != None:
                                webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
                                for row in webhookcsv:
                                    if (x == row[0]):
                                        webhookurlrow =  row
                                        webhookurl = webhookurlrow[1]
                                        result = False
                                        session = requests.session()
                                        data = {"Content": "/md \n\n### Scheduler on break \nHello team, **" + NameScheduler + "** is going for a break. Please note that @" + loginbuddy2 + " will be your routing POC in the meantime." }
                                        print(data)
                                        params = {'format': 'application/json'}
                                        response = session.post(webhookurl, params=params, json=data)
                                        if response.status_code == 200:
                                            result = True
                            else:
                                print("error")
                        
                        for x in DSscheduler3coverbuddycountry4:       
                            if x != None:
                                webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
                                for row in webhookcsv:
                                    if (x == row[0]):
                                        webhookurlrow =  row
                                        webhookurl = webhookurlrow[1]
                                        result = False
                                        session = requests.session()
                                        data = {"Content": "/md \n\n### Scheduler on break \nHello team, **" + NameScheduler + "** is going for a break. Please note that @" + loginbuddy3 + " will be your routing POC in the meantime." }
                                        print(data)
                                        params = {'format': 'application/json'}
                                        response = session.post(webhookurl, params=params, json=data)
                                        if response.status_code == 200:
                                            result = True
                            else:
                                print("error")
            
            tkinter.messagebox.showinfo('Stations Handed Over', 'Webhook sent to all chime rooms \n Enjoy your break!!')
        
        #back from break
        def backfrombreak():
            os.chdir(r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\Chime Broadcaster')#webhook addresses
            answer = tkinter.messagebox.askokcancel("IMPORTANT", "By clicking OK you will be sending Webhooks to all your DS notifying you will are back from your break")
            if answer:
                if Buddy3 == None and Buddy4==None:
                    Sched1 = Buddy1list[1] #take the 2nd element from the buddylist
                    Sched2 = Buddy2list[1] #do the same
                    
                    if NameScheduler== Sched1:
                        for x in DSscheduler1coverbuddycountry1:       
                            if x != None:
                                webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
                                for row in webhookcsv:
                                    if (x == row[0]):
                                        webhookurlrow =  row
                                        webhookurl = webhookurlrow[1]
                                        result = False
                                        session = requests.session()
                                        data = {"Content": "/md \n\n### Scheduler back from break \nHello team, **" + NameScheduler + "** is now back from the break" }
                                        print(data)
                                        params = {'format': 'application/json'}
                                        response = session.post(webhookurl, params=params, json=data)
                                        if response.status_code == 200:
                                            result = True

                            else:
                                print("error")
                    

                    if NameScheduler== Sched2:
                        for x in DSscheduler1coverbuddycountry2:       
                            if x != None:
                                webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
                                for row in webhookcsv:
                                    if (x == row[0]):
                                        webhookurlrow =  row
                                        webhookurl = webhookurlrow[1]
                                        result = False
                                        session = requests.session()
                                        data = {"Content": "/md \n\n### Scheduler back from break \nHello team, **" + NameScheduler + "** is now back from the break" }
                                        print(data)
                                        params = {'format': 'application/json'}
                                        response = session.post(webhookurl, params=params, json=data)
                                        if response.status_code == 200:
                                            result = True

                            else:
                                print("error")
                        
                


                if Buddy3 != None and Buddy4 == None:
                    Sched1 = Buddy1list[2] #take the 3rd element from the buddylist to get just 1 name
                    Sched2 = Buddy2list[2]
                    Sched3 = Buddy3list[2]
                    
                    if NameScheduler == Sched1:
                        for x in DSscheduler1coverbuddycountry1:       
                            if x != None:
                                webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
                                for row in webhookcsv:
                                    if (x == row[0]):
                                        webhookurlrow =  row
                                        webhookurl = webhookurlrow[1]
                                        result = False
                                        session = requests.session()
                                        data = {"Content": "/md \n\n### Scheduler back from break \nHello team, **" + NameScheduler + "** is now back from the break" }
                                        print(data)
                                        params = {'format': 'application/json'}
                                        response = session.post(webhookurl, params=params, json=data)
                                        if response.status_code == 200:
                                            result = True

                            else:
                                print("error")
                    
                        for x in DSscheduler2coverbuddycountry1:       
                            if x != None:
                                webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
                                for row in webhookcsv:
                                    if (x == row[0]):
                                        webhookurlrow =  row
                                        webhookurl = webhookurlrow[1]
                                        result = False
                                        session = requests.session()
                                        data = {"Content": "/md \n\n### Scheduler back from break \nHello team, **" + NameScheduler + "** is now back from the break" }
                                        print(data)
                                        params = {'format': 'application/json'}
                                        response = session.post(webhookurl, params=params, json=data)
                                        if response.status_code == 200:
                                            result = True
                            
                            else:
                                print("error")


                    if NameScheduler== Sched2:
                        for x in DSscheduler1coverbuddycountry2:       
                            if x != None:
                                webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
                                for row in webhookcsv:
                                    if (x == row[0]):
                                        webhookurlrow =  row
                                        webhookurl = webhookurlrow[1]
                                        result = False
                                        session = requests.session()
                                        data = {"Content": "/md \n\n### Scheduler back from break \nHello team, **" + NameScheduler + "** is now back from the break" }
                                        print(data)
                                        params = {'format': 'application/json'}
                                        response = session.post(webhookurl, params=params, json=data)
                                        if response.status_code == 200:
                                            result = True
                            
                            else:
                                print("error")
                        for x in DSscheduler2coverbuddycountry2:       
                            if x != None:
                                webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
                                for row in webhookcsv:
                                    if (x == row[0]):
                                        webhookurlrow =  row
                                        webhookurl = webhookurlrow[1]
                                        print(x + webhookurl)
                                        result = False
                                        session = requests.session()
                                        data = {"Content": "/md \n\n### Scheduler back from break \nHello team, **" + NameScheduler + "** is now back from the break" }
                                        print(data)
                                        params = {'format': 'application/json'}
                                        response = session.post(webhookurl, params=params, json=data)
                                        if response.status_code == 200:
                                            result = True
                            else:
                                print("error")
                        
                        
                    if NameScheduler== Sched3:
                        for x in DSscheduler1coverbuddycountry3:       
                            if x != None:
                                webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
                                for row in webhookcsv:
                                    if (x == row[0]):
                                        webhookurlrow =  row
                                        webhookurl = webhookurlrow[1]
                                        result = False
                                        session = requests.session()
                                        data = {"Content": "/md \n\n### Scheduler back from break \nHello team, **" + NameScheduler + "** is now back from the break" }
                                        print(data)
                                        params = {'format': 'application/json'}
                                        response = session.post(webhookurl, params=params, json=data)
                                        if response.status_code == 200:
                                            result = True
                            else:
                                print("error")
                        for x in DSscheduler2coverbuddycountry3:       
                            if x != None:
                                webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
                                for row in webhookcsv:
                                    if (x == row[0]):
                                        webhookurlrow =  row
                                        webhookurl = webhookurlrow[1]
                                        result = False
                                        session = requests.session()
                                        data = {"Content": "/md \n\n### Scheduler back from break \nHello team, **" + NameScheduler + "** is now back from the break" }
                                        print(data)
                                        params = {'format': 'application/json'}
                                        response = session.post(webhookurl, params=params, json=data)
                                        if response.status_code == 200:
                                            result = True

                            else:
                                print("error")
                        
                
                if Buddy4 != None:
                    Sched1 = Buddy1list[3] #take the 4th element from the buddylist to get just 1 name
                    Sched2 = Buddy2list[3]
                    Sched3 = Buddy3list[3]
                    Sched4 = Buddy4list[3]
                    
                    if NameScheduler== Sched1:
                        for x in DSscheduler1coverbuddycountry1:       
                            if x != None:
                                webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
                                for row in webhookcsv:
                                    if (x == row[0]):
                                        webhookurlrow =  row
                                        webhookurl = webhookurlrow[1]
                                        result = False
                                        session = requests.session()
                                        data = {"Content": "/md \n\n### Scheduler back from break \nHello team, **" + NameScheduler + "** is now back from the break" }
                                        print(data)
                                        params = {'format': 'application/json'}
                                        response = session.post(webhookurl, params=params, json=data)
                                        if response.status_code == 200:
                                            result = True
                            else:
                                print("error")
                        for x in DSscheduler2coverbuddycountry1:       
                            if x != None:
                                webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
                                for row in webhookcsv:
                                    if (x == row[0]):
                                        webhookurlrow =  row
                                        webhookurl = webhookurlrow[1]
                                        result = False
                                        session = requests.session()
                                        data = {"Content": "/md \n\n### Scheduler back from break \nHello team, **" + NameScheduler + "** is now back from the break" }
                                        print(data)
                                        params = {'format': 'application/json'}
                                        response = session.post(webhookurl, params=params, json=data)
                                        if response.status_code == 200:
                                            result = True
                            else:
                                print("error")
                        
                        for x in DSscheduler3coverbuddycountry1:       
                            if x != None:
                                webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
                                for row in webhookcsv:
                                    if (x == row[0]):
                                        webhookurlrow =  row
                                        webhookurl = webhookurlrow[1]
                                        result = False
                                        session = requests.session()
                                        data = {"Content": "/md \n\n### Scheduler back from break \nHello team, **" + NameScheduler + "** is now back from the break" }
                                        print(data)
                                        params = {'format': 'application/json'}
                                        response = session.post(webhookurl, params=params, json=data)
                                        if response.status_code == 200:
                                            result = True
                            else:
                                print("error")
                        
                        
                        
                        
                        
                    
                    if NameScheduler== Sched2:
                        for x in DSscheduler1coverbuddycountry2:       
                            if x != None:
                                webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
                                for row in webhookcsv:
                                    if (x == row[0]):
                                        webhookurlrow =  row
                                        webhookurl = webhookurlrow[1]
                                        result = False
                                        session = requests.session()
                                        data = {"Content": "/md \n\n### Scheduler back from break \nHello team, **" + NameScheduler + "** is now back from the break" }
                                        print(data)
                                        params = {'format': 'application/json'}
                                        response = session.post(webhookurl, params=params, json=data)
                                        if response.status_code == 200:
                                            result = True
                            else:
                                print("error")
                        
                        for x in DSscheduler2coverbuddycountry2:       
                            if x != None:
                                webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
                                for row in webhookcsv:
                                    if (x == row[0]):
                                        webhookurlrow =  row
                                        webhookurl = webhookurlrow[1]
                                        result = False
                                        session = requests.session()
                                        data = {"Content": "/md \n\n### Scheduler back from break \nHello team, **" + NameScheduler + "** is now back from the break" }
                                        print(data)
                                        params = {'format': 'application/json'}
                                        response = session.post(webhookurl, params=params, json=data)
                                        if response.status_code == 200:
                                            result = True
                            else:
                                print("error")
                        
                        for x in DSscheduler3coverbuddycountry2:       
                            if x != None:
                                webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
                                for row in webhookcsv:
                                    if (x == row[0]):
                                        webhookurlrow =  row
                                        webhookurl = webhookurlrow[1]
                                        result = False
                                        session = requests.session()
                                        data = {"Content": "/md \n\n### Scheduler back from break \nHello team, **" + NameScheduler + "** is now back from the break" }
                                        print(data)
                                        params = {'format': 'application/json'}
                                        response = session.post(webhookurl, params=params, json=data)
                                        if response.status_code == 200:
                                            result = True
                            else:
                                print("error")
                        
                    
                        

                    if NameScheduler== Sched3:
                        for x in DSscheduler1coverbuddycountry3:       
                            if x != None:
                                webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
                                for row in webhookcsv:
                                    if (x == row[0]):
                                        webhookurlrow =  row
                                        webhookurl = webhookurlrow[1]
                                        result = False
                                        session = requests.session()
                                        data = {"Content": "/md \n\n### Scheduler back from break \nHello team, **" + NameScheduler + "** is now back from the break" }
                                        print(data)
                                        params = {'format': 'application/json'}
                                        response = session.post(webhookurl, params=params, json=data)
                                        if response.status_code == 200:
                                            result = True
                            else:
                                print("error")
                        
                        for x in DSscheduler2coverbuddycountry3:       
                            if x != None:
                                webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
                                for row in webhookcsv:
                                    if (x == row[0]):
                                        webhookurlrow =  row
                                        webhookurl = webhookurlrow[1]
                                        result = False
                                        session = requests.session()
                                        data = {"Content": "/md \n\n### Scheduler back from break \nHello team, **" + NameScheduler + "** is now back from the break" }
                                        print(data)
                                        params = {'format': 'application/json'}
                                        response = session.post(webhookurl, params=params, json=data)
                                        if response.status_code == 200:
                                            result = True
                            else:
                                print("error")
                        
                        for x in DSscheduler3coverbuddycountry3:       
                            if x != None:
                                webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
                                for row in webhookcsv:
                                    if (x == row[0]):
                                        webhookurlrow =  row
                                        webhookurl = webhookurlrow[1]
                                        result = False
                                        session = requests.session()
                                        data = {"Content": "/md \n\n### Scheduler back from break \nHello team, **" + NameScheduler + "** is now back from the break" }
                                        print(data)
                                        params = {'format': 'application/json'}
                                        response = session.post(webhookurl, params=params, json=data)
                                        if response.status_code == 200:
                                            result = True
                            else:
                                print("error")

                        
                        
                        

                    if NameScheduler== Sched4:
                        for x in DSscheduler1coverbuddycountry4:       
                            if x != None:
                                webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
                                for row in webhookcsv:
                                    if (x == row[0]):
                                        webhookurlrow =  row
                                        webhookurl = webhookurlrow[1]
                                        result = False
                                        session = requests.session()
                                        data = {"Content": "/md \n\n### Scheduler back from break \nHello team, **" + NameScheduler + "** is now back from the break" }
                                        print(data)
                                        params = {'format': 'application/json'}
                                        response = session.post(webhookurl, params=params, json=data)
                                        if response.status_code == 200:
                                            result = True
                            else:
                                print("error")
                        
                        for x in DSscheduler2coverbuddycountry4:       
                            if x != None:
                                webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
                                for row in webhookcsv:
                                    if (x == row[0]):
                                        webhookurlrow =  row
                                        webhookurl = webhookurlrow[1]
                                        result = False
                                        session = requests.session()
                                        data = {"Content": "/md \n\n### Scheduler back from break \nHello team, **" + NameScheduler + "** is now back from the break" }
                                        print(data)
                                        params = {'format': 'application/json'}
                                        response = session.post(webhookurl, params=params, json=data)
                                        if response.status_code == 200:
                                            result = True
                            else:
                                print("error")
                        
                        for x in DSscheduler3coverbuddycountry4:       
                            if x != None:
                                webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
                                for row in webhookcsv:
                                    if (x == row[0]):
                                        webhookurlrow =  row
                                        webhookurl = webhookurlrow[1]
                                        result = False
                                        session = requests.session()
                                        data = {"Content": "/md \n\n### Scheduler back from break \nHello team, **" + NameScheduler + "** is now back from the break" }
                                        print(data)
                                        params = {'format': 'application/json'}
                                        response = session.post(webhookurl, params=params, json=data)
                                        if response.status_code == 200:
                                            result = True
                            else:
                                print("error")
            
            tkinter.messagebox.showinfo('Break finished', 'Webhook sent to all chime rooms \n You are now officially back on shift')


        

        def updatetime():
            top = Toplevel()
            top.title("Check at what time the break slot was selected")
            os.chdir(r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\TaskList\TasklistGeneration\\' + dateStr2 )
            try:
                breakcsv = csv.reader(open("breakinfo.csv","r"))
                for row in breakcsv:
                    if search_strbuddy1 == row[0]:
                        Breakinfo = row
                        timeupdatescheduler1 = Breakinfo[2]
                        print(Breakinfo)
                        print(timeupdatescheduler1)
            except:
                pass

            try:
                breakcsv = csv.reader(open("breakinfo.csv","r"))
                for row in breakcsv:
                    if search_strbuddy2 == row[0]:
                        Breakinfo2 = row
                        timeupdatescheduler2 = Breakinfo2[2]           
            except:
                pass

            try:
                breakcsv = csv.reader(open("breakinfo.csv","r"))
                for row in breakcsv:
                    if search_strbuddy3 == row[0]:
                        Breakinfo3 = row
                        timeupdatescheduler3 = Breakinfo3[2]           
            except:
                pass

            try:
                breakcsv = csv.reader(open("breakinfo.csv","r"))
                for row in breakcsv:
                    if search_strbuddy4 == row[0]:
                        Breakinfo4 = row
                        timeupdatescheduler4 = Breakinfo4[2]           
            except:
                pass

            
                
            try:
                if timeupdatescheduler1 != None:
                    timeupdatescheduler1label = Label(top, text= search_strbuddy1 + " selected breaks at " + str(timeupdatescheduler1), foreground= "black", relief=GROOVE, justify=CENTER, padx=20, pady=5)
                    timeupdatescheduler1label.grid(column = 0, row= 1, pady=10, padx=10, columnspan=3)
            except:
                pass
            try:
                if timeupdatescheduler2 != None:
                    timeupdatescheduler2label = Label(top, text= search_strbuddy2 + " selected breaks at " + str(timeupdatescheduler2), foreground= "black", relief=GROOVE, justify=CENTER, padx=20, pady=5)
                    timeupdatescheduler2label.grid(column = 0, row= 2, pady=10, padx=10, columnspan=3)
            except:
                pass
            try:
                if timeupdatescheduler3 != None:
                    timeupdatescheduler3label = Label(top, text= search_strbuddy3 + " selected breaks at " + str(timeupdatescheduler3), foreground= "black", relief=GROOVE, justify=CENTER, padx=20, pady=5)
                    timeupdatescheduler3label.grid(column = 0, row= 3, pady=10, padx=10, columnspan=3)
            except:
                pass
            try:
                if timeupdatescheduler4 != None:
                    timeupdatescheduler4label = Label(top, text= search_strbuddy4 + " selected breaks at " + str(timeupdatescheduler4), foreground= "black", relief=GROOVE, justify=CENTER, padx=20, pady=5)
                    timeupdatescheduler4label.grid(column = 0, row= 4, pady=10, padx=10, columnspan=3)
            except:
                pass


        

        # Information = Label(frame.scrollable_frame, text= "If any extra cover, inform buddies", justify = CENTER,fg="black", font=("Helvetica", 10, 'bold'))
        # Information.grid(row=7, column=3, columnspan=3, pady=5, padx=10)

        def handover():
            top = Toplevel()
            top.title("Station Broadcaster")
            Startingbreak = Button(top, text = "GOING FOR BREAK", fg= "red", command = reminder, width=30)
            Startingbreak.grid(row=2, column=0, columnspan=1, pady=30, padx=40)
            Backfrombreak = Button(top, text = "COMING FROM BREAK", fg= "red", command = backfrombreak, width=30)
            Backfrombreak.grid(row=4, column=0, columnspan=1, pady=30, padx=40)
            
        
        Reminder = Button(frame.scrollable_frame, text = "Station Broadcaster", fg= "red", command = handover, width=15)
        Reminder.grid(row=0, column=8, columnspan=1, pady=10, padx=10)
        
        Updatetime = Button(frame.scrollable_frame, text = "Who picked first?", command = updatetime, width=15)
        Updatetime.grid(row=0, column=9, columnspan=1, pady=10, padx=10)


        
        


    def extracover():
            global DS1
            DS1 = DS1entry.get()
            DS1entry.delete(0,END)
            DS1cover= DS1.upper()

            global DS2
            DS2 = DS2entry.get()
            DS2entry.delete(0,END)
            DS2cover= DS2.upper()
            
            try: 
                os.chdir(r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\TaskList\TasklistGeneration\\' + dateStr2 )
                f = open("extracover.csv","a", newline="")
                coverextra = (NameScheduler, DS1cover, DS2cover)
                writer = csv.writer(f)
                writer.writerow(coverextra)
                f.close()
        
            except Exception as FILEOPEN:
                print("File appears to be open. Please close the file and try again!")
                pass

    frameextracover = ttk.LabelFrame(frame.scrollable_frame, padding=5)
    frameextracover.grid(row=4, column = 0, columnspan=4, \
    padx=5, pady=5, ipadx=5, ipady=5)

    Addbutton = Button(frameextracover, text = "Add", command=lambda:[extracover(), refresh()] , width=15)
    Addbutton.grid(row=4, column=2, columnspan=3, pady=10, padx=10) 


    def confirm():
        question = tkinter.messagebox.askyesno('Confirm covering','Are you happy covering your buddies on the time slots displayed?')
        if question == TRUE:
            try: 
                os.chdir(r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\TaskList\TasklistGeneration\\' + dateStr2 )
                f = open("breakinfoconfirmation.csv","a", newline="")
                tup = (NameScheduler, "Confirmed")
                writer = csv.writer(f)
                writer.writerow(tup)
                f.close()
        
            except Exception as FILEOPEN:
                print("File appears to be open. Please close the file and try again!")
                pass

        else:
            pass  




    Confirm = Button(frame.scrollable_frame, text = "CONFIRM", state = DISABLED, command=lambda:[confirm(), refresh()], width=15)
    Confirm.grid(row=0, column=10, columnspan=1, pady=10, padx=10) 
            
    framesearch = ttk.LabelFrame(frame.scrollable_frame, padding=5)
    framesearch.grid(row=0, column = 0, columnspan=5, \
    padx=5, pady=5, ipadx=5, ipady=5)

    Timeentry = Entry(framesearch, width=30, borderwidth=5)
    Timeentry.grid(row=0, column=1, pady=10, padx=10, columnspan=3)

    Updatetimeentry = Button(framesearch, text = "Update", command=lambda:[update(Timeentry), refresh()], width=15)
    Updatetimeentry.grid(row=0, column=4, columnspan=1, pady=10, padx=10) 


    DS1label = Label(frameextracover, text= "DS extra cover", justify = CENTER,fg="black", font=("Helvetica", 10, 'bold'))
    DS1label.grid(row=4, column=0, columnspan=1, pady=5, padx=10)
    DS2label = Label(frameextracover, text= "DS extra cover", justify = CENTER,fg="black", font=("Helvetica", 10, 'bold'))
    DS2label.grid(row=5, column=0, columnspan=1, pady=5, padx=10)

    DS1entry = Entry(frameextracover, width=15, borderwidth=5)
    DS1entry.grid(row=4, column=1, columnspan=1, pady=10, padx=10) 
    DS2entry = Entry(frameextracover, width=15, borderwidth=5)
    DS2entry.grid(row=5, column=1, columnspan=1, pady=10, padx=10) 


    Refreshbutton = Button(frame.scrollable_frame, text = "Refresh", command = refresh, width=15)
    Refreshbutton.grid(row=0, column=5, columnspan=3, pady=10, padx=10)
    Refreshbutton.invoke()


    textlabelbreak1 = ttk.Label(framesearch,  text= "Break time for " + user_login + "\n (hh:mm to hh:mm)", justify=LEFT, padding=5).grid(row=0, column=0, columnspan=1, pady=10, padx=10)

    



    frame.pack(fill = tk.BOTH, expand = True)

    root.mainloop()

if __name__ == '__main__':
    SplitBuddiesExecutable()